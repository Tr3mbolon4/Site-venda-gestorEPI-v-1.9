from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Request
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from database import connect_db, close_db, get_db
from schemas import *
from auth import *
from datetime import datetime, timezone, timedelta
from typing import List, Optional
from pathlib import Path
from bson import ObjectId
import os
import shutil
import logging
import base64
import json
import io
import asyncio
import resend

# Para importação de Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Para geração de PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).parent
UPLOAD_DIR = ROOT_DIR / 'uploads'
UPLOAD_DIR.mkdir(exist_ok=True)

# Dias para expiração de senha
PASSWORD_EXPIRY_DAYS = 30

# Resend Email Config
resend.api_key = os.environ.get('RESEND_API_KEY')
CONTACT_EMAIL = os.environ.get('CONTACT_EMAIL', 'alexandre_santos@prismaxshop.com.br')

app = FastAPI(title='GestorEPI API')
api_router = APIRouter(prefix='/api')

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=['*'],
    allow_methods=['*'],
    allow_headers=['*'],
    expose_headers=['*'],
)

# Middleware para adicionar headers de cache-control e CORS em arquivos estáticos
@app.middleware("http")
async def add_cache_headers(request, call_next):
    response = await call_next(request)
    
    # Para arquivos de upload, adicionar headers de cache e CORS
    if '/uploads/' in request.url.path or '/api/uploads/' in request.url.path:
        response.headers['Cache-Control'] = 'public, max-age=86400'  # 1 dia
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        response.headers['X-Content-Type-Options'] = 'nosniff'
    
    return response

# Servir arquivos de upload - DEVE ter prefixo /api para funcionar com ingress
app.mount('/api/uploads', StaticFiles(directory=str(UPLOAD_DIR)), name='api_uploads')
app.mount('/uploads', StaticFiles(directory=str(UPLOAD_DIR)), name='uploads')

@app.on_event('startup')
async def startup_event():
    await connect_db()
    from seed import seed_database
    await seed_database()

@app.on_event('shutdown')
async def shutdown_event():
    await close_db()

def doc_to_response(doc, id_field='id'):
    if doc is None:
        return None
    result = {k: v for k, v in doc.items() if k != '_id'}
    result[id_field] = str(doc['_id'])
    return result

def check_password_expired(user):
    """Verifica se a senha expirou (mais de 30 dias)"""
    password_changed_at = user.get('password_changed_at')
    if not password_changed_at:
        return False
    
    if password_changed_at.tzinfo is None:
        password_changed_at = password_changed_at.replace(tzinfo=timezone.utc)
    
    expiry_date = password_changed_at + timedelta(days=PASSWORD_EXPIRY_DAYS)
    return datetime.now(timezone.utc) > expiry_date

# Permissões por perfil
ROLE_PERMISSIONS = {
    'admin': ['all'],
    'gestor': ['dashboard', 'entrega', 'colaboradores', 'empresas', 'epis', 'fornecedores', 'kits'],
    'rh': ['dashboard', 'colaboradores', 'colaboradores_full', 'empresas', 'usuarios'],
    'seguranca_trabalho': ['dashboard', 'epis', 'fornecedores', 'kits', 'colaboradores_list'],
    'almoxarifado': ['dashboard', 'entrega', 'colaboradores_list']
}

def can_view_sensitive_data(role):
    """Verifica se o perfil pode ver dados sensíveis (CPF, RG, etc)"""
    return role in ['admin', 'gestor', 'rh']

def can_manage_users(role):
    """Verifica se pode gerenciar usuários"""
    return role in ['admin', 'rh']

def can_deliver_epi(role):
    """Verifica se pode fazer entregas de EPI"""
    return role in ['admin', 'gestor', 'almoxarifado']

def can_manage_epis(role):
    """Verifica se pode gerenciar EPIs - inclui almoxarifado"""
    return role in ['admin', 'gestor', 'seguranca_trabalho', 'almoxarifado']

def can_manage_employees(role):
    """Verifica se pode cadastrar/editar colaboradores"""
    return role in ['admin', 'gestor', 'rh']

# ===================== AUTH =====================

@api_router.get('/')
async def root():
    return {'message': 'GestorEPI API'}

@api_router.get('/health')
async def health_check():
    """Endpoint de verificação de saúde da API"""
    # Verificar conexão com MongoDB
    db_status = 'unknown'
    try:
        db = await get_db()
        await db.command('ping')
        db_status = 'connected'
    except Exception as e:
        db_status = f'error: {str(e)}'
    
    # Verificar pasta de uploads
    uploads_status = 'ok' if os.path.isdir(UPLOAD_DIR) else 'missing'
    
    # Contar arquivos de upload
    upload_count = 0
    if os.path.isdir(UPLOAD_DIR):
        for root_dir, dirs, files in os.walk(UPLOAD_DIR):
            upload_count += len(files)
    
    return {
        'status': 'healthy' if db_status == 'connected' else 'degraded',
        'database': db_status,
        'uploads_dir': uploads_status,
        'upload_files_count': upload_count,
        'backend_url': os.environ.get('BACKEND_URL', 'not_set'),
        'timestamp': datetime.now(timezone.utc).isoformat()
    }

# ===================== CONTACT FORM (EMAIL) =====================

class ContactFormRequest(BaseModel):
    nome: str
    empresa: str
    telefone: str
    mensagem: Optional[str] = None

@api_router.post('/contact')
async def send_contact_form(data: ContactFormRequest):
    """Envia email do formulário de contato do site"""
    if not resend.api_key:
        logger.error("RESEND_API_KEY não configurada")
        raise HTTPException(status_code=500, detail="Serviço de email não configurado")
    
    html_content = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background: linear-gradient(135deg, #f97316, #ea580c); padding: 20px; border-radius: 10px 10px 0 0;">
            <h1 style="color: white; margin: 0; font-size: 24px;">Nova Solicitação - GestorEPI</h1>
        </div>
        <div style="background: #f9f9f9; padding: 30px; border: 1px solid #e0e0e0; border-top: none; border-radius: 0 0 10px 10px;">
            <h2 style="color: #333; margin-top: 0;">Dados do Contato</h2>
            <table style="width: 100%; border-collapse: collapse;">
                <tr>
                    <td style="padding: 10px 0; border-bottom: 1px solid #eee; font-weight: bold; color: #555;">Nome:</td>
                    <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #333;">{data.nome}</td>
                </tr>
                <tr>
                    <td style="padding: 10px 0; border-bottom: 1px solid #eee; font-weight: bold; color: #555;">Empresa:</td>
                    <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #333;">{data.empresa}</td>
                </tr>
                <tr>
                    <td style="padding: 10px 0; border-bottom: 1px solid #eee; font-weight: bold; color: #555;">Telefone:</td>
                    <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #333;">{data.telefone}</td>
                </tr>
                <tr>
                    <td style="padding: 10px 0; font-weight: bold; color: #555; vertical-align: top;">Mensagem:</td>
                    <td style="padding: 10px 0; color: #333;">{data.mensagem or 'Não informada'}</td>
                </tr>
            </table>
            <div style="margin-top: 30px; padding: 15px; background: #fff3e0; border-radius: 8px; border-left: 4px solid #f97316;">
                <p style="margin: 0; color: #666; font-size: 14px;">
                    <strong>Próximo passo:</strong> Entre em contato com o cliente o mais breve possível.
                </p>
            </div>
        </div>
        <div style="text-align: center; padding: 20px; color: #888; font-size: 12px;">
            <p>Email enviado automaticamente pelo site GestorEPI</p>
        </div>
    </div>
    """
    
    params = {
        "from": "GestorEPI <onboarding@resend.dev>",
        "to": [CONTACT_EMAIL],
        "subject": f"Nova Solicitação: {data.empresa} - {data.nome}",
        "html": html_content
    }
    
    try:
        email = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Email enviado com sucesso para {CONTACT_EMAIL} - ID: {email.get('id')}")
        return {
            "status": "success",
            "message": "Mensagem enviada com sucesso! Entraremos em contato em breve.",
            "email_id": email.get("id")
        }
    except Exception as e:
        logger.error(f"Erro ao enviar email: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao enviar mensagem: {str(e)}")

@api_router.post('/auth/login', response_model=TokenResponse)
async def login(request: LoginRequest):
    db = await get_db()
    
    # Verificar se o identificador é email ou username
    identifier = request.username.strip()
    if '@' in identifier:
        # Buscar por email
        user = await db.users.find_one({"email": identifier.lower()})
    else:
        # Buscar por username
        user = await db.users.find_one({"username": identifier})
    
    if not user or not verify_password(request.password, user['hashed_password']):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail='Credenciais inválidas')
    
    if not user.get('is_active', True):
        raise HTTPException(status_code=400, detail='Usuário inativo')
    
    # MULTI-TENANT: Verificar empresa do usuário
    empresa_id = user.get('empresa_id')
    empresa_nome = None
    
    if empresa_id and user.get('role') != 'super_admin':
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
        if empresa:
            if empresa.get('status') == 'bloqueado':
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail='Empresa bloqueada. Entre em contato com o administrador.'
                )
            empresa_nome = empresa.get('nome')
    
    license_doc = await db.panel_license.find_one({})
    if license_doc:
        now = datetime.now(timezone.utc)
        expires_at = license_doc['expires_at']
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=timezone.utc)
        if now > expires_at and user['role'] not in ['admin', 'super_admin']:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail='Licença expirada')
    
    password_expired = check_password_expired(user)
    
    access_token = create_access_token(data={'sub': user['username'], 'role': user['role']})
    
    return TokenResponse(
        access_token=access_token,
        token_type='bearer',
        must_change_password=user.get('must_change_password', False),
        password_expired=password_expired,
        role=UserRole(user['role']),
        is_primary_admin=user.get('is_primary_admin', False),
        empresa_id=empresa_id,
        empresa_nome=empresa_nome
    )

@api_router.post('/auth/change-password')
async def change_password(request: ChangePasswordRequest, current_user: dict = Depends(get_current_user)):
    db = await get_db()
    if not verify_password(request.old_password, current_user['hashed_password']):
        raise HTTPException(status_code=400, detail='Senha antiga incorreta')
    
    await db.users.update_one(
        {"_id": ObjectId(current_user['id'])},
        {"$set": {
            "hashed_password": get_password_hash(request.new_password), 
            "must_change_password": False,
            "password_changed_at": datetime.now(timezone.utc)
        }}
    )
    return {'message': 'Senha alterada com sucesso'}

@api_router.get('/auth/me', response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(
        id=current_user['id'],
        username=current_user['username'],
        email=current_user['email'],
        role=UserRole(current_user['role']),
        is_active=current_user.get('is_active', True),
        must_change_password=current_user.get('must_change_password', False),
        password_changed_at=current_user.get('password_changed_at'),
        employee_id=current_user.get('employee_id'),
        empresa_id=current_user.get('empresa_id'),
        empresa_nome=current_user.get('empresa_nome'),
        created_at=current_user.get('created_at', datetime.now(timezone.utc))
    )

# ===================== EMPRESAS (MULTI-TENANT) - SUPER_ADMIN =====================

@api_router.get('/empresas', response_model=List[EmpresaResponse])
async def get_empresas(current_user: dict = Depends(require_super_admin())):
    """Lista todas as empresas (apenas SUPER_ADMIN)"""
    db = await get_db()
    empresas = await db.empresas.find({}).to_list(1000)
    
    result = []
    for emp in empresas:
        # Contar colaboradores
        colab_count = await db.employees.count_documents({"empresa_id": str(emp['_id'])})
        resp = doc_to_response(emp)
        resp['colaboradores_cadastrados'] = colab_count
        result.append(EmpresaResponse(**resp))
    
    return result

@api_router.post('/empresas', response_model=EmpresaResponse, status_code=status.HTTP_201_CREATED)
async def create_empresa(empresa_data: EmpresaCreate, current_user: dict = Depends(require_super_admin())):
    """Cria nova empresa (apenas SUPER_ADMIN)"""
    db = await get_db()
    
    # Verificar CNPJ único
    existing = await db.empresas.find_one({"cnpj": empresa_data.cnpj})
    if existing:
        raise HTTPException(status_code=400, detail='CNPJ já cadastrado')
    
    # Definir limite baseado no plano
    plano_limites = {
        "50": 50, "150": 150, "250": 250, "350": 350, "unlimited": 999999
    }
    limite = plano_limites.get(empresa_data.plano.value, 50)
    
    new_empresa = {
        "nome": empresa_data.nome,
        "cnpj": empresa_data.cnpj,
        "status": empresa_data.status.value,
        "plano": empresa_data.plano.value,
        "limite_colaboradores": empresa_data.limite_colaboradores or limite,
        "endereco": empresa_data.endereco,
        "telefone": empresa_data.telefone,
        "email": empresa_data.email,
        "responsavel": empresa_data.responsavel,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    result = await db.empresas.insert_one(new_empresa)
    new_empresa['_id'] = result.inserted_id
    
    resp = doc_to_response(new_empresa)
    resp['colaboradores_cadastrados'] = 0
    return EmpresaResponse(**resp)

@api_router.get('/empresas/{empresa_id}', response_model=EmpresaResponse)
async def get_empresa(empresa_id: str, current_user: dict = Depends(require_super_admin())):
    """Busca empresa por ID (apenas SUPER_ADMIN)"""
    db = await get_db()
    
    empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    if not empresa:
        raise HTTPException(status_code=404, detail='Empresa não encontrada')
    
    colab_count = await db.employees.count_documents({"empresa_id": empresa_id})
    resp = doc_to_response(empresa)
    resp['colaboradores_cadastrados'] = colab_count
    return EmpresaResponse(**resp)

@api_router.patch('/empresas/{empresa_id}', response_model=EmpresaResponse)
async def update_empresa(empresa_id: str, empresa_data: EmpresaUpdate, current_user: dict = Depends(require_super_admin())):
    """Atualiza empresa (apenas SUPER_ADMIN)"""
    db = await get_db()
    
    update_data = {k: v for k, v in empresa_data.model_dump(exclude_unset=True).items() if v is not None}
    
    # Converter enums para string
    if 'status' in update_data:
        update_data['status'] = update_data['status'].value
    if 'plano' in update_data:
        update_data['plano'] = update_data['plano'].value
    
    update_data['updated_at'] = datetime.now(timezone.utc)
    
    result = await db.empresas.find_one_and_update(
        {"_id": ObjectId(empresa_id)},
        {"$set": update_data},
        return_document=True
    )
    
    if not result:
        raise HTTPException(status_code=404, detail='Empresa não encontrada')
    
    colab_count = await db.employees.count_documents({"empresa_id": empresa_id})
    resp = doc_to_response(result)
    resp['colaboradores_cadastrados'] = colab_count
    return EmpresaResponse(**resp)

@api_router.post('/empresas/{empresa_id}/bloquear')
async def bloquear_empresa(empresa_id: str, current_user: dict = Depends(require_super_admin())):
    """Bloqueia uma empresa (apenas SUPER_ADMIN)"""
    db = await get_db()
    
    result = await db.empresas.find_one_and_update(
        {"_id": ObjectId(empresa_id)},
        {"$set": {"status": "bloqueado", "updated_at": datetime.now(timezone.utc)}},
        return_document=True
    )
    
    if not result:
        raise HTTPException(status_code=404, detail='Empresa não encontrada')
    
    return {'message': f'Empresa {result["nome"]} bloqueada com sucesso'}

@api_router.post('/empresas/{empresa_id}/ativar')
async def ativar_empresa(empresa_id: str, current_user: dict = Depends(require_super_admin())):
    """Ativa uma empresa bloqueada (apenas SUPER_ADMIN)"""
    db = await get_db()
    
    result = await db.empresas.find_one_and_update(
        {"_id": ObjectId(empresa_id)},
        {"$set": {"status": "ativo", "updated_at": datetime.now(timezone.utc)}},
        return_document=True
    )
    
    if not result:
        raise HTTPException(status_code=404, detail='Empresa não encontrada')
    
    return {'message': f'Empresa {result["nome"]} ativada com sucesso'}

@api_router.post('/empresas/{empresa_id}/criar-admin')
async def criar_admin_empresa(empresa_id: str, user_data: UserCreate, current_user: dict = Depends(require_super_admin())):
    """Cria usuário admin para uma empresa (apenas SUPER_ADMIN)"""
    db = await get_db()
    
    # Verificar se empresa existe
    empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    if not empresa:
        raise HTTPException(status_code=404, detail='Empresa não encontrada')
    
    # Verificar se username/email já existe
    email_normalized = user_data.email.lower().strip()
    existing = await db.users.find_one({"$or": [
        {"username": user_data.username},
        {"email": email_normalized}
    ]})
    if existing:
        raise HTTPException(status_code=400, detail='Usuário ou e-mail já existe')
    
    new_user = {
        "username": user_data.username,
        "email": email_normalized,
        "hashed_password": get_password_hash(user_data.password),
        "role": "admin",
        "empresa_id": empresa_id,
        "must_change_password": True,
        "is_active": True,
        "password_changed_at": datetime.now(timezone.utc),
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    result = await db.users.insert_one(new_user)
    new_user['_id'] = result.inserted_id
    
    resp = doc_to_response(new_user)
    resp['empresa_nome'] = empresa['nome']
    return UserResponse(**resp)

@api_router.get('/empresas/{empresa_id}/stats')
async def get_empresa_stats(empresa_id: str, current_user: dict = Depends(require_super_admin())):
    """Estatísticas de uso de uma empresa (apenas SUPER_ADMIN)"""
    db = await get_db()
    
    empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    if not empresa:
        raise HTTPException(status_code=404, detail='Empresa não encontrada')
    
    # Contar registros
    colaboradores = await db.employees.count_documents({"empresa_id": empresa_id})
    usuarios = await db.users.count_documents({"empresa_id": empresa_id})
    epis = await db.epis.count_documents({"empresa_id": empresa_id})
    entregas = await db.deliveries.count_documents({"empresa_id": empresa_id})
    kits = await db.kits.count_documents({"empresa_id": empresa_id})
    
    return {
        'empresa_id': empresa_id,
        'empresa_nome': empresa['nome'],
        'plano': empresa.get('plano'),
        'limite_colaboradores': empresa.get('limite_colaboradores'),
        'colaboradores': colaboradores,
        'usuarios': usuarios,
        'epis': epis,
        'entregas': entregas,
        'kits': kits,
        'uso_percentual': round((colaboradores / empresa.get('limite_colaboradores', 50)) * 100, 1) if empresa.get('limite_colaboradores') else 0
    }

# ===================== USERS =====================

@api_router.get('/users', response_model=List[UserResponse])
async def get_users(current_user: dict = Depends(require_role('admin', 'rh'))):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    empresa_filter = get_empresa_filter(current_user)
    users = await db.users.find(empresa_filter).to_list(1000)
    
    result = []
    for u in users:
        resp = doc_to_response(u)
        # Buscar nome da empresa
        if u.get('empresa_id'):
            emp = await db.empresas.find_one({"_id": ObjectId(u['empresa_id'])})
            resp['empresa_nome'] = emp['nome'] if emp else None
        result.append(UserResponse(**resp))
    
    return result

@api_router.post('/users', response_model=UserResponse, status_code=status.HTTP_201_CREATED)
async def create_user(user_data: UserCreate, current_user: dict = Depends(require_role('admin', 'rh'))):
    db = await get_db()
    
    # RH não pode criar admins
    if current_user['role'] == 'rh' and user_data.role == UserRole.ADMIN:
        raise HTTPException(status_code=403, detail='Sem permissão para criar administradores')
    
    # Normalizar email para lowercase
    email_normalized = user_data.email.lower().strip() if user_data.email else None
    
    existing = await db.users.find_one({"$or": [
        {"username": user_data.username}, 
        {"email": email_normalized}
    ]})
    if existing:
        if existing.get('username') == user_data.username:
            raise HTTPException(status_code=400, detail='Nome de usuário já existe')
        else:
            raise HTTPException(status_code=400, detail='E-mail já está em uso')
    
    # MULTI-TENANT: Associar usuário à mesma empresa do criador
    empresa_id = current_user.get('empresa_id')
    if current_user.get('role') != 'super_admin' and not empresa_id:
        raise HTTPException(status_code=400, detail='Usuário deve estar associado a uma empresa')
    
    new_user = {
        "username": user_data.username,
        "email": email_normalized,
        "hashed_password": get_password_hash(user_data.password),
        "role": user_data.role.value,
        "employee_id": user_data.employee_id,
        "empresa_id": empresa_id,  # MULTI-TENANT
        "must_change_password": True,
        "is_active": True,
        "password_changed_at": datetime.now(timezone.utc),
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    result = await db.users.insert_one(new_user)
    new_user['_id'] = result.inserted_id
    
    resp = doc_to_response(new_user)
    resp['empresa_nome'] = current_user.get('empresa_nome')
    return UserResponse(**resp)

@api_router.patch('/users/{user_id}', response_model=UserResponse)
async def update_user(user_id: str, user_data: UserUpdate, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    update_data = {k: v for k, v in user_data.model_dump(exclude_unset=True).items()}
    if 'role' in update_data:
        update_data['role'] = update_data['role'].value
    update_data['updated_at'] = datetime.now(timezone.utc)
    
    result = await db.users.find_one_and_update(
        {"_id": ObjectId(user_id)}, {"$set": update_data}, return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail='Usuário não encontrado')
    return UserResponse(**doc_to_response(result))

@api_router.delete('/users/{user_id}')
async def delete_user(user_id: str, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    result = await db.users.delete_one({"_id": ObjectId(user_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Usuário não encontrado')
    return {'message': 'Usuário excluído'}

class ResetPasswordRequest(BaseModel):
    new_password: str

@api_router.post('/users/{user_id}/reset-password')
async def reset_user_password(user_id: str, request: ResetPasswordRequest, current_user: dict = Depends(require_role('admin'))):
    """Permite ao administrador redefinir a senha de qualquer usuário"""
    db = await get_db()
    
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail='Usuário não encontrado')
    
    # Validar complexidade da senha
    password = request.new_password
    import re
    if len(password) < 8:
        raise HTTPException(status_code=400, detail='A senha deve ter no mínimo 8 caracteres')
    if not re.search(r'[A-Z]', password):
        raise HTTPException(status_code=400, detail='A senha deve conter pelo menos uma letra maiúscula')
    if not re.search(r'[a-z]', password):
        raise HTTPException(status_code=400, detail='A senha deve conter pelo menos uma letra minúscula')
    if not re.search(r'\d', password):
        raise HTTPException(status_code=400, detail='A senha deve conter pelo menos um número')
    if not re.search(r'[!@#$%^&*(),.?":{}|<>]', password):
        raise HTTPException(status_code=400, detail='A senha deve conter pelo menos um caractere especial')
    
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {
            "hashed_password": get_password_hash(password),
            "must_change_password": True,
            "password_changed_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {'message': 'Senha redefinida com sucesso'}

# ===================== COMPANIES =====================

@api_router.get('/companies', response_model=List[CompanyResponse])
async def get_companies(current_user: dict = Depends(require_role('admin', 'gestor', 'rh'))):
    db = await get_db()
    companies = await db.companies.find({}).to_list(1000)
    return [CompanyResponse(**doc_to_response(c)) for c in companies]

@api_router.post('/companies', response_model=CompanyResponse, status_code=status.HTTP_201_CREATED)
async def create_company(company_data: CompanyCreate, current_user: dict = Depends(require_role('admin', 'gestor', 'rh'))):
    logger.info(f"📝 Recebendo dados de empresa: {company_data.model_dump()}")
    db = await get_db()
    existing = await db.companies.find_one({"cnpj": company_data.cnpj})
    if existing:
        logger.error(f"❌ CNPJ já cadastrado: {company_data.cnpj}")
        raise HTTPException(status_code=400, detail='CNPJ já cadastrado')
    
    new_company = {**company_data.model_dump(), "created_at": datetime.now(timezone.utc), "updated_at": datetime.now(timezone.utc)}
    logger.info(f"✅ Inserindo empresa: {new_company}")
    result = await db.companies.insert_one(new_company)
    new_company['_id'] = result.inserted_id
    logger.info(f"✅ Empresa criada com ID: {result.inserted_id}")
    return CompanyResponse(**doc_to_response(new_company))

@api_router.get('/companies/{company_id}', response_model=CompanyResponse)
async def get_company(company_id: str, current_user: dict = Depends(require_role('admin', 'gestor', 'rh'))):
    db = await get_db()
    company = await db.companies.find_one({"_id": ObjectId(company_id)})
    if not company:
        raise HTTPException(status_code=404, detail='Empresa não encontrada')
    return CompanyResponse(**doc_to_response(company))

@api_router.patch('/companies/{company_id}', response_model=CompanyResponse)
async def update_company(company_id: str, company_data: CompanyUpdate, current_user: dict = Depends(require_role('admin', 'gestor', 'rh'))):
    db = await get_db()
    update_data = {k: v for k, v in company_data.model_dump(exclude_unset=True).items()}
    update_data['updated_at'] = datetime.now(timezone.utc)
    result = await db.companies.find_one_and_update(
        {"_id": ObjectId(company_id)}, {"$set": update_data}, return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail='Empresa não encontrada')
    return CompanyResponse(**doc_to_response(result))

@api_router.delete('/companies/{company_id}')
async def delete_company(company_id: str, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    result = await db.companies.delete_one({"_id": ObjectId(company_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Empresa não encontrada')
    return {'message': 'Empresa excluída'}

# ===================== EMPLOYEES =====================

@api_router.get('/employees')
async def get_employees(current_user: dict = Depends(get_current_user), search: Optional[str] = None, company_id: Optional[str] = None):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = get_empresa_filter(current_user)
    
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"cpf": {"$regex": search, "$options": "i"}},
            {"registration_number": {"$regex": search, "$options": "i"}}
        ]
    if company_id:
        query["company_id"] = company_id
    
    employees = await db.employees.find(query).to_list(1000)
    
    # Perfis que não podem ver dados sensíveis
    if not can_view_sensitive_data(current_user['role']):
        return [EmployeePublicResponse(**doc_to_response(e)) for e in employees]
    
    return [EmployeeResponse(**doc_to_response(e)) for e in employees]

@api_router.post('/employees', response_model=EmployeeResponse, status_code=status.HTTP_201_CREATED)
async def create_employee(employee_data: EmployeeCreate, current_user: dict = Depends(get_current_user)):
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissão para cadastrar colaboradores')
    
    db = await get_db()
    
    # MULTI-TENANT: Verificar limite de colaboradores
    empresa_id = current_user.get('empresa_id')
    if empresa_id:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
        if empresa:
            limite = empresa.get('limite_colaboradores', 50)
            atual = await db.employees.count_documents({"empresa_id": empresa_id})
            if atual >= limite:
                raise HTTPException(
                    status_code=400, 
                    detail=f'Limite de colaboradores atingido ({atual}/{limite}). Atualize seu plano.'
                )
    
    # Verificar CPF único na mesma empresa
    cpf_query = {"cpf": employee_data.cpf}
    if empresa_id:
        cpf_query["empresa_id"] = empresa_id
    existing = await db.employees.find_one(cpf_query)
    if existing:
        raise HTTPException(status_code=400, detail='CPF já cadastrado nesta empresa')
    
    new_employee = {
        **employee_data.model_dump(), 
        "empresa_id": empresa_id,  # MULTI-TENANT
        "created_at": datetime.now(timezone.utc), 
        "updated_at": datetime.now(timezone.utc)
    }
    result = await db.employees.insert_one(new_employee)
    new_employee['_id'] = result.inserted_id
    return EmployeeResponse(**doc_to_response(new_employee))

@api_router.get('/employees/{employee_id}')
async def get_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = {"_id": ObjectId(employee_id)}
    empresa_filter = get_empresa_filter(current_user)
    if empresa_filter:
        query.update(empresa_filter)
    
    employee = await db.employees.find_one(query)
    if not employee:
        raise HTTPException(status_code=404, detail='Colaborador não encontrado')
    
    if not can_view_sensitive_data(current_user['role']):
        return EmployeePublicResponse(**doc_to_response(employee))
    
    return EmployeeResponse(**doc_to_response(employee))

@api_router.patch('/employees/{employee_id}', response_model=EmployeeResponse)
async def update_employee(employee_id: str, employee_data: EmployeeUpdate, current_user: dict = Depends(get_current_user)):
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissão para editar colaboradores')
    
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = {"_id": ObjectId(employee_id)}
    empresa_filter = get_empresa_filter(current_user)
    if empresa_filter:
        query.update(empresa_filter)
    
    update_data = {k: v for k, v in employee_data.model_dump(exclude_unset=True).items()}
    update_data['updated_at'] = datetime.now(timezone.utc)
    result = await db.employees.find_one_and_update(
        query, {"$set": update_data}, return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail='Colaborador não encontrado')
    return EmployeeResponse(**doc_to_response(result))

@api_router.delete('/employees/{employee_id}')
async def delete_employee(employee_id: str, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = {"_id": ObjectId(employee_id)}
    empresa_filter = get_empresa_filter(current_user)
    if empresa_filter:
        query.update(empresa_filter)
    
    result = await db.employees.delete_one(query)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Colaborador não encontrado')
    return {'message': 'Colaborador excluído'}

@api_router.post('/employees/{employee_id}/photo')
async def upload_employee_photo(employee_id: str, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissão')
    
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = {"_id": ObjectId(employee_id)}
    empresa_filter = get_empresa_filter(current_user)
    if empresa_filter:
        query.update(empresa_filter)
    
    employee = await db.employees.find_one(query)
    if not employee:
        raise HTTPException(status_code=404, detail='Colaborador não encontrado')
    
    file_ext = Path(file.filename).suffix
    file_name = f'employee_{employee_id}_{datetime.now(timezone.utc).timestamp()}{file_ext}'
    file_path = UPLOAD_DIR / 'employees' / file_name
    file_path.parent.mkdir(exist_ok=True, parents=True)
    
    with file_path.open('wb') as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    photo_path = f'/uploads/employees/{file_name}'
    await db.employees.update_one(query, {"$set": {"photo_path": photo_path}})
    return {'photo_path': photo_path}

# ===================== IMPORTAÇÃO/EXPORTAÇÃO =====================

@api_router.get('/employees/export/excel')
async def export_employees_excel(current_user: dict = Depends(get_current_user)):
    """Exporta colaboradores para Excel"""
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Permissão insuficiente')
    
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = get_empresa_filter(current_user)
    
    employees = await db.employees.find(query).to_list(5000)
    companies = {str(c['_id']): c['legal_name'] async for c in db.companies.find()}
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Colaboradores"
    
    # Cabeçalho
    headers = ['Nome Completo', 'CPF', 'RG', 'Matrícula', 'Empresa', 'Cargo', 'Setor', 'Status']
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Dados
    for row, emp in enumerate(employees, 2):
        ws.cell(row=row, column=1, value=emp.get('full_name', ''))
        ws.cell(row=row, column=2, value=emp.get('cpf', ''))
        ws.cell(row=row, column=3, value=emp.get('rg', ''))
        ws.cell(row=row, column=4, value=emp.get('registration_number', ''))
        ws.cell(row=row, column=5, value=companies.get(emp.get('company_id', ''), ''))
        ws.cell(row=row, column=6, value=emp.get('position', ''))
        ws.cell(row=row, column=7, value=emp.get('department', ''))
        ws.cell(row=row, column=8, value='Ativo' if emp.get('status') == 'active' else 'Inativo')
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=colaboradores_{datetime.now().strftime("%Y%m%d")}.xlsx'}
    )

@api_router.get('/employees/template/excel')
async def download_employees_template(current_user: dict = Depends(get_current_user)):
    """Download do template Excel para importação de colaboradores"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Colaboradores"
    
    # Cabeçalho com instruções
    headers = ['Nome Completo*', 'CPF*', 'RG', 'Matrícula*', 'Empresa*', 'Cargo', 'Setor', 'Status']
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Exemplo de preenchimento
    ws.cell(row=2, column=1, value='João da Silva')
    ws.cell(row=2, column=2, value='123.456.789-00')
    ws.cell(row=2, column=3, value='12.345.678-9')
    ws.cell(row=2, column=4, value='MAT001')
    ws.cell(row=2, column=5, value='Nome da Empresa')
    ws.cell(row=2, column=6, value='Operador')
    ws.cell(row=2, column=7, value='Produção')
    ws.cell(row=2, column=8, value='Ativo')
    
    # Instruções
    ws2 = wb.create_sheet(title="Instruções")
    ws2.cell(row=1, column=1, value="INSTRUÇÕES DE PREENCHIMENTO").font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value="* Campos obrigatórios")
    ws2.cell(row=4, column=1, value="• Nome Completo: Nome completo do colaborador")
    ws2.cell(row=5, column=1, value="• CPF: Formato XXX.XXX.XXX-XX ou apenas números")
    ws2.cell(row=6, column=1, value="• Matrícula: Código único do colaborador na empresa")
    ws2.cell(row=7, column=1, value="• Empresa: Nome exato da empresa cadastrada no sistema")
    ws2.cell(row=8, column=1, value="• Status: 'Ativo' ou 'Inativo' (padrão: Ativo)")
    
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=template_colaboradores.xlsx'}
    )

@api_router.post('/employees/import/excel')
async def import_employees_excel(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Importa colaboradores de um arquivo Excel"""
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Permissão insuficiente')
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail='Arquivo deve ser Excel (.xlsx ou .xls)')
    
    db = await get_db()
    
    # Carregar empresas
    companies = {}
    async for company in db.companies.find():
        companies[company['legal_name'].lower().strip()] = str(company['_id'])
    
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        
        results = {'imported': 0, 'errors': [], 'skipped': 0}
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:  # Linha vazia
                continue
            
            full_name = str(row[0]).strip() if row[0] else None
            cpf = str(row[1]).strip() if row[1] else None
            rg = str(row[2]).strip() if row[2] else None
            registration_number = str(row[3]).strip() if row[3] else None
            company_name = str(row[4]).strip().lower() if row[4] else None
            position = str(row[5]).strip() if len(row) > 5 and row[5] else None
            department = str(row[6]).strip() if len(row) > 6 and row[6] else None
            status_str = str(row[7]).strip().lower() if len(row) > 7 and row[7] else 'ativo'
            
            # Validações
            errors = []
            if not full_name:
                errors.append('Nome é obrigatório')
            if not cpf:
                errors.append('CPF é obrigatório')
            if not registration_number:
                errors.append('Matrícula é obrigatória')
            if not company_name:
                errors.append('Empresa é obrigatória')
            elif company_name not in companies:
                errors.append(f'Empresa "{row[4]}" não encontrada no sistema')
            
            if errors:
                results['errors'].append({'row': row_num, 'errors': errors, 'name': full_name})
                continue
            
            # Verificar se já existe
            existing = await db.employees.find_one({
                "$or": [
                    {"cpf": cpf},
                    {"registration_number": registration_number, "company_id": companies.get(company_name)}
                ]
            })
            
            if existing:
                results['skipped'] += 1
                results['errors'].append({
                    'row': row_num, 
                    'errors': ['Colaborador já existe (CPF ou Matrícula duplicada)'],
                    'name': full_name
                })
                continue
            
            # Inserir colaborador
            employee = {
                'full_name': full_name,
                'cpf': cpf,
                'rg': rg,
                'registration_number': registration_number,
                'company_id': companies.get(company_name),
                'position': position,
                'department': department,
                'status': 'active' if status_str == 'ativo' else 'inactive',
                'facial_consent': False,
                'created_at': datetime.now(timezone.utc),
                'updated_at': datetime.now(timezone.utc)
            }
            
            await db.employees.insert_one(employee)
            results['imported'] += 1
        
        return results
        
    except Exception as e:
        logger.error(f"Erro ao importar Excel: {str(e)}")
        raise HTTPException(status_code=400, detail=f'Erro ao processar arquivo: {str(e)}')

# ===================== IMPRESSÃO PDF =====================

@api_router.get('/reports/employees/pdf')
async def generate_employees_pdf(current_user: dict = Depends(get_current_user)):
    """Gera relatório PDF de colaboradores"""
    if not can_view_sensitive_data(current_user['role']):
        raise HTTPException(status_code=403, detail='Permissão insuficiente')
    
    db = await get_db()
    employees = await db.employees.find({"status": "active"}).to_list(5000)
    companies = {str(c['_id']): c['legal_name'] async for c in db.companies.find()}
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=20)
    
    elements = []
    
    # Título
    elements.append(Paragraph("Relatório de Colaboradores", title_style))
    elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Tabela
    data = [['Nome', 'CPF', 'Matrícula', 'Empresa', 'Cargo', 'Setor']]
    
    for emp in employees:
        data.append([
            emp.get('full_name', '')[:30],
            emp.get('cpf', ''),
            emp.get('registration_number', ''),
            companies.get(emp.get('company_id', ''), '')[:25],
            (emp.get('position', '') or '')[:20],
            (emp.get('department', '') or '')[:15]
        ])
    
    table = Table(data, colWidths=[120, 90, 70, 120, 100, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.063, 0.725, 0.506)),  # Emerald
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Total de colaboradores ativos: {len(employees)}", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=colaboradores_{datetime.now().strftime("%Y%m%d")}.pdf'}
    )

@api_router.get('/reports/deliveries/pdf')
async def generate_deliveries_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Gera relatório PDF de entregas"""
    db = await get_db()
    
    query = {}
    if start_date:
        query['created_at'] = {'$gte': datetime.fromisoformat(start_date.replace('Z', '+00:00'))}
    if end_date:
        if 'created_at' not in query:
            query['created_at'] = {}
        query['created_at']['$lte'] = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    
    deliveries = await db.deliveries.find(query).sort('created_at', -1).to_list(1000)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=20)
    
    elements = []
    
    # Título
    elements.append(Paragraph("Relatório de Entregas de EPIs", title_style))
    elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Tabela
    data = [['Data', 'Colaborador', 'Itens', 'Tipo', 'Verificação Facial']]
    
    for delivery in deliveries:
        items_str = ', '.join([
            f"{item.get('quantity', 1)}x {item.get('epi_name', item.get('kit_name', 'Item'))}"
            for item in delivery.get('items', [])
        ])[:50]
        
        facial_match = delivery.get('facial_match_score')
        facial_str = f"{int(facial_match * 100)}%" if facial_match else 'N/A'
        
        data.append([
            delivery.get('created_at', datetime.now()).strftime('%d/%m/%Y %H:%M'),
            delivery.get('employee_name', '')[:25],
            items_str,
            'Devolução' if delivery.get('is_return') else 'Entrega',
            facial_str
        ])
    
    table = Table(data, colWidths=[100, 150, 200, 70, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.063, 0.725, 0.506)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Total de entregas: {len(deliveries)}", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=entregas_{datetime.now().strftime("%Y%m%d")}.pdf'}
    )

@api_router.get('/reports/employee/{employee_id}/pdf')
async def generate_employee_history_pdf(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Gera PDF com ficha do colaborador e histórico de entregas"""
    db = await get_db()
    
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail='Colaborador não encontrado')
    
    company = None
    if employee.get('company_id'):
        company = await db.companies.find_one({"_id": ObjectId(employee['company_id'])})
    
    deliveries = await db.deliveries.find({"employee_id": employee_id}).sort('created_at', -1).to_list(500)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=40)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, spaceAfter=10)
    
    elements = []
    
    # Cabeçalho
    elements.append(Paragraph("Ficha do Colaborador", title_style))
    elements.append(Spacer(1, 10))
    
    # Dados do colaborador
    info_data = [
        ['Nome:', employee.get('full_name', '')],
        ['CPF:', employee.get('cpf', '')],
        ['RG:', employee.get('rg', '') or '-'],
        ['Matrícula:', employee.get('registration_number', '')],
        ['Empresa:', company.get('legal_name', '') if company else '-'],
        ['Cargo:', employee.get('position', '') or '-'],
        ['Setor:', employee.get('department', '') or '-'],
        ['Status:', 'Ativo' if employee.get('status') == 'active' else 'Inativo'],
    ]
    
    info_table = Table(info_data, colWidths=[100, 350])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 30))
    
    # Histórico de entregas
    elements.append(Paragraph("Histórico de Entregas/Devoluções", subtitle_style))
    
    if deliveries:
        # ATUALIZADO: Incluir responsável pela entrega
        history_data = [['Data', 'Tipo', 'Itens', 'Responsável', 'Verificação']]
        
        for delivery in deliveries:
            items_str = ', '.join([
                f"{item.get('quantity', 1)}x {item.get('epi_name', item.get('kit_name', 'Item'))}"
                for item in delivery.get('items', [])
            ])[:50]
            
            facial_match = delivery.get('facial_match_score')
            facial_str = f"{int(facial_match * 100)}%" if facial_match else '-'
            
            # Nome do responsável pela entrega
            responsavel = delivery.get('delivered_by_name', '-')
            
            history_data.append([
                delivery.get('created_at', datetime.now()).strftime('%d/%m/%Y'),
                'Devolução' if delivery.get('is_return') else 'Entrega',
                items_str,
                responsavel[:20],  # Limitar tamanho do nome
                facial_str
            ])
        
        history_table = Table(history_data, colWidths=[65, 60, 180, 100, 55])
        history_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.063, 0.725, 0.506)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
        ]))
        
        elements.append(history_table)
    else:
        elements.append(Paragraph("Nenhuma entrega registrada.", styles['Normal']))
    
    elements.append(Spacer(1, 40))
    elements.append(Paragraph(f"Documento gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=ficha_{employee.get("registration_number", employee_id)}_{datetime.now().strftime("%Y%m%d")}.pdf'}
    )

# ===================== FACIAL TEMPLATES =====================

@api_router.get('/employees/{employee_id}/facial-templates')
async def get_facial_templates(employee_id: str, current_user: dict = Depends(get_current_user)):
    db = await get_db()
    templates = await db.facial_templates.find({"employee_id": employee_id}).to_list(100)
    return [doc_to_response(t) for t in templates]

# Endpoint otimizado para buscar TODOS os templates de uma vez
@api_router.get('/facial-templates/all')
async def get_all_facial_templates(current_user: dict = Depends(get_current_user)):
    """Retorna todos os templates faciais com informações do colaborador - otimizado para reconhecimento
    MULTI-TENANT: Retorna apenas templates de colaboradores da mesma empresa
    """
    db = await get_db()
    
    # MULTI-TENANT: Filtrar colaboradores por empresa
    empresa_filter = get_empresa_filter(current_user)
    emp_query = {}
    if empresa_filter:
        emp_query.update(empresa_filter)
    
    # Buscar colaboradores da empresa primeiro
    employees = await db.employees.find(emp_query).to_list(5000)
    if not employees:
        return []
    
    # Criar mapa de colaboradores e lista de IDs
    emp_map = {}
    employee_ids = []
    for e in employees:
        emp_id = str(e['_id'])
        emp_map[emp_id] = doc_to_response(e)
        employee_ids.append(ObjectId(e['_id']))
    
    # Buscar templates APENAS dos colaboradores da empresa
    templates = await db.facial_templates.find({
        "employee_id": {"$in": [str(eid) for eid in employee_ids] + employee_ids}
    }).to_list(5000)
    
    if not templates:
        return []
    
    # Montar resposta com dados do colaborador
    result = []
    for t in templates:
        emp_id = t.get('employee_id')
        emp_id_str = str(emp_id) if emp_id else None
        if emp_id_str and emp_id_str in emp_map:
            result.append({
                'id': str(t['_id']),
                'employee_id': emp_id_str,
                'descriptor': t.get('descriptor'),
                'employee': emp_map[emp_id_str]
            })
    
    return result

# ===================== VERIFICAÇÃO DE DUPLICIDADE BIOMÉTRICA =====================

BIOMETRIC_DUPLICATE_THRESHOLD = 0.40  # Distância máxima para considerar duplicata

def euclidean_distance(a, b):
    """Calcula distância euclidiana entre dois vetores"""
    import math
    return math.sqrt(sum((x - y) ** 2 for x, y in zip(a, b)))

@api_router.post('/biometric/check-duplicate')
async def check_biometric_duplicate(
    data: BiometricCheckRequest,
    current_user: dict = Depends(get_current_user)
):
    """Verifica se existe biometria duplicada no sistema - MULTI-TENANT: apenas na mesma empresa"""
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    empresa_filter = get_empresa_filter(current_user)
    
    try:
        # Parsear o descriptor enviado
        new_descriptor = json.loads(data.descriptor)
        if len(new_descriptor) != 128:
            raise HTTPException(status_code=400, detail="Descriptor deve ter 128 valores")
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Descriptor inválido")
    
    # Buscar colaboradores da empresa primeiro
    emp_query = {}
    if empresa_filter:
        emp_query.update(empresa_filter)
    employees = await db.employees.find(emp_query).to_list(5000)
    employee_ids = [str(e['_id']) for e in employees]
    emp_map = {str(e['_id']): e for e in employees}
    
    # Buscar templates apenas dos colaboradores da empresa
    templates = await db.facial_templates.find({
        "employee_id": {"$in": employee_ids}
    }).to_list(1000)
    
    # Comparar com cada template
    for template in templates:
        emp_id = str(template.get('employee_id'))
        
        # Ignorar o próprio colaborador (para edição)
        if data.employee_id and emp_id == data.employee_id:
            continue
        
        try:
            existing_descriptor = json.loads(template.get('descriptor', '[]'))
            if len(existing_descriptor) != 128:
                continue
            
            # Calcular distância
            distance = euclidean_distance(new_descriptor, existing_descriptor)
            similarity = 1 - distance
            
            # Se a distância for menor que o threshold, é duplicata
            if distance < BIOMETRIC_DUPLICATE_THRESHOLD:
                emp_data = emp_map.get(emp_id, {})
                return BiometricCheckResponse(
                    is_duplicate=True,
                    duplicate_employee_id=emp_id,
                    duplicate_employee_name=emp_data.get('full_name', 'Desconhecido'),
                    similarity_score=similarity,
                    message=f"Esta foto já corresponde ou está muito semelhante à biometria facial do colaborador '{emp_data.get('full_name', 'Desconhecido')}'. Não é permitido utilizar a mesma foto para colaboradores diferentes."
                )
        except Exception as e:
            continue
    
    return BiometricCheckResponse(
        is_duplicate=False,
        message="Biometria válida - nenhuma duplicata encontrada"
    )

@api_router.get('/biometric/audit-duplicates')
async def audit_biometric_duplicates(current_user: dict = Depends(get_current_user)):
    """Audita a base de dados para encontrar biometrias duplicadas existentes"""
    if current_user['role'] not in ['admin', 'gestor']:
        raise HTTPException(status_code=403, detail="Apenas administradores podem auditar duplicatas")
    
    db = await get_db()
    
    # Buscar todos os templates
    templates = await db.facial_templates.find({}).to_list(1000)
    
    # Buscar colaboradores
    employee_ids = list(set([t.get('employee_id') for t in templates if t.get('employee_id')]))
    employees = await db.employees.find({"_id": {"$in": [ObjectId(str(e)) for e in employee_ids]}}).to_list(1000)
    emp_map = {str(e['_id']): e for e in employees}
    
    duplicates = []
    checked_pairs = set()
    
    # Comparar todos os pares
    for i, t1 in enumerate(templates):
        emp_id1 = str(t1.get('employee_id'))
        
        try:
            desc1 = json.loads(t1.get('descriptor', '[]'))
            if len(desc1) != 128:
                continue
        except:
            continue
        
        for j, t2 in enumerate(templates):
            if i >= j:
                continue
            
            emp_id2 = str(t2.get('employee_id'))
            
            # Não comparar mesmo colaborador
            if emp_id1 == emp_id2:
                continue
            
            # Evitar duplicar pares
            pair_key = tuple(sorted([emp_id1, emp_id2]))
            if pair_key in checked_pairs:
                continue
            checked_pairs.add(pair_key)
            
            try:
                desc2 = json.loads(t2.get('descriptor', '[]'))
                if len(desc2) != 128:
                    continue
                
                distance = euclidean_distance(desc1, desc2)
                
                if distance < BIOMETRIC_DUPLICATE_THRESHOLD:
                    emp1 = emp_map.get(emp_id1, {})
                    emp2 = emp_map.get(emp_id2, {})
                    
                    duplicates.append({
                        "employee1_id": emp_id1,
                        "employee1_name": emp1.get('full_name', 'Desconhecido'),
                        "employee2_id": emp_id2,
                        "employee2_name": emp2.get('full_name', 'Desconhecido'),
                        "similarity_score": round(1 - distance, 4),
                        "distance": round(distance, 4)
                    })
            except:
                continue
    
    return {
        "total_templates": len(templates),
        "duplicates_found": len(duplicates),
        "duplicates": duplicates
    }

@api_router.post('/employees/{employee_id}/biometric-consent')
async def register_biometric_consent(
    employee_id: str,
    consent_data: BiometricConsentRequest,
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Registra o consentimento biométrico do colaborador (LGPD)"""
    db = await get_db()
    
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    
    # Obter IP do cliente
    client_ip = consent_data.ip_address or request.client.host
    
    # Atualizar colaborador com dados de consentimento
    update_data = {
        "facial_consent": consent_data.accepted,
        "facial_consent_date": datetime.now(timezone.utc),
        "facial_consent_ip": client_ip
    }
    
    await db.employees.update_one(
        {"_id": ObjectId(employee_id)},
        {"$set": update_data}
    )
    
    # Registrar log de consentimento para auditoria
    consent_log = {
        "employee_id": ObjectId(employee_id),
        "employee_name": employee.get('full_name'),
        "accepted": consent_data.accepted,
        "ip_address": client_ip,
        "user_agent": request.headers.get("user-agent", "unknown"),
        "registered_by": current_user.get("username"),
        "created_at": datetime.now(timezone.utc)
    }
    await db.biometric_consent_logs.insert_one(consent_log)
    
    return {
        "success": True,
        "message": "Consentimento biométrico registrado com sucesso",
        "consent_date": update_data["facial_consent_date"].isoformat(),
        "ip_address": client_ip
    }

@api_router.get('/employees/{employee_id}/biometric-consent')
async def get_biometric_consent(
    employee_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Verifica o status do consentimento biométrico do colaborador"""
    db = await get_db()
    
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    
    return {
        "has_consent": employee.get("facial_consent", False),
        "consent_date": employee.get("facial_consent_date"),
        "consent_ip": employee.get("facial_consent_ip")
    }

@api_router.post('/employees/{employee_id}/facial-templates')
async def create_facial_template(employee_id: str, template_data: FacialTemplateCreate, current_user: dict = Depends(get_current_user)):
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissão')
    
    db = await get_db()
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail='Colaborador não encontrado')
    
    new_template = {
        "employee_id": ObjectId(employee_id),
        "descriptor": template_data.descriptor,
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.facial_templates.insert_one(new_template)
    
    # Retornar resposta formatada corretamente
    return {
        "id": str(result.inserted_id),
        "employee_id": employee_id,
        "descriptor": template_data.descriptor,
        "created_at": new_template["created_at"].isoformat()
    }

@api_router.delete('/employees/{employee_id}/facial-templates/{template_id}')
async def delete_facial_template(employee_id: str, template_id: str, current_user: dict = Depends(get_current_user)):
    if not can_manage_employees(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissão')
    
    db = await get_db()
    result = await db.facial_templates.delete_one({
        "_id": ObjectId(template_id),
        "employee_id": ObjectId(employee_id)
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Template não encontrado')
    
    return {'message': 'Template excluído'}

# ===================== SUPPLIERS =====================

@api_router.get('/suppliers', response_model=List[SupplierResponse])
async def get_suppliers(current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    # MULTI-TENANT: Filtrar por empresa
    query = get_empresa_filter(current_user)
    suppliers = await db.suppliers.find(query).to_list(1000)
    return [SupplierResponse(**doc_to_response(s)) for s in suppliers]

@api_router.post('/suppliers', response_model=SupplierResponse, status_code=status.HTTP_201_CREATED)
async def create_supplier(supplier_data: SupplierCreate, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    # MULTI-TENANT: Associar à empresa do usuário
    empresa_id = current_user.get('empresa_id')
    
    # Validar CNPJ duplicado DENTRO DA MESMA EMPRESA
    if supplier_data.cnpj:
        existing = await db.suppliers.find_one({
            "cnpj": supplier_data.cnpj,
            "empresa_id": empresa_id
        })
        if existing:
            raise HTTPException(status_code=400, detail='CNPJ já cadastrado nesta empresa')
    
    new_supplier = {
        **supplier_data.model_dump(), 
        "empresa_id": empresa_id,  # MULTI-TENANT
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.suppliers.insert_one(new_supplier)
    new_supplier['_id'] = result.inserted_id
    return SupplierResponse(**doc_to_response(new_supplier))

@api_router.patch('/suppliers/{supplier_id}', response_model=SupplierResponse)
async def update_supplier(supplier_id: str, supplier_data: SupplierUpdate, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    # MULTI-TENANT: Filtrar por empresa
    query = {"_id": ObjectId(supplier_id)}
    empresa_filter = get_empresa_filter(current_user)
    if empresa_filter:
        query.update(empresa_filter)
    
    update_data = {k: v for k, v in supplier_data.model_dump(exclude_unset=True).items()}
    update_data['updated_at'] = datetime.now(timezone.utc)
    result = await db.suppliers.find_one_and_update(
        query, {"$set": update_data}, return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail='Fornecedor não encontrado')
    return SupplierResponse(**doc_to_response(result))

@api_router.delete('/suppliers/{supplier_id}')
async def delete_supplier(supplier_id: str, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    # MULTI-TENANT: Filtrar por empresa
    query = {"_id": ObjectId(supplier_id)}
    empresa_filter = get_empresa_filter(current_user)
    if empresa_filter:
        query.update(empresa_filter)
    
    result = await db.suppliers.delete_one(query)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Fornecedor não encontrado')
    return {'message': 'Fornecedor excluído'}

# ===================== EPIS =====================

def calculate_epi_status(epi):
    """Calcula status de estoque e validade do EPI"""
    stock_status = 'ok'
    validity_status = 'ok'
    
    # Status de estoque
    if epi.get('current_stock', 0) <= 0:
        stock_status = 'out'
    elif epi.get('current_stock', 0) <= epi.get('min_stock', 0):
        stock_status = 'low'
    
    # Status de validade
    validity_date = epi.get('validity_date') or epi.get('ca_validity')
    if validity_date:
        if validity_date.tzinfo is None:
            validity_date = validity_date.replace(tzinfo=timezone.utc)
        
        now = datetime.now(timezone.utc)
        days_until_expiry = (validity_date - now).days
        
        if days_until_expiry < 0:
            validity_status = 'expired'
        elif days_until_expiry <= 30:
            validity_status = 'expiring'
    
    return stock_status, validity_status

@api_router.get('/epis', response_model=List[EPIResponse])
async def get_epis(current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho', 'almoxarifado'))):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = get_empresa_filter(current_user)
    epis = await db.epis.find(query).to_list(1000)
    
    result = []
    for e in epis:
        stock_status, validity_status = calculate_epi_status(e)
        resp = doc_to_response(e)
        resp['stock_status'] = stock_status
        resp['validity_status'] = validity_status
        result.append(EPIResponse(**resp))
    
    return result

@api_router.post('/epis', response_model=EPIResponse, status_code=status.HTTP_201_CREATED)
async def create_epi(epi_data: EPICreate, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    
    # Validar que pelo menos CA ou NBR deve ser preenchido
    if not epi_data.ca_number and not epi_data.nbr_number:
        raise HTTPException(status_code=400, detail='É necessário informar o número do CA ou NBR')
    
    # MULTI-TENANT: Associar à empresa do usuário
    empresa_id = current_user.get('empresa_id')
    
    new_epi = {
        **epi_data.model_dump(), 
        "empresa_id": empresa_id,  # MULTI-TENANT
        "created_by": current_user['id'], 
        "created_at": datetime.now(timezone.utc), 
        "updated_at": datetime.now(timezone.utc)
    }
    
    # Converter enum para string se necessário
    if new_epi.get('replacement_period'):
        new_epi['replacement_period'] = str(new_epi['replacement_period'].value) if hasattr(new_epi['replacement_period'], 'value') else str(new_epi['replacement_period'])
    
    result = await db.epis.insert_one(new_epi)
    new_epi['_id'] = result.inserted_id
    stock_status, validity_status = calculate_epi_status(new_epi)
    resp = doc_to_response(new_epi)
    resp['stock_status'] = stock_status
    resp['validity_status'] = validity_status
    return EPIResponse(**resp)

@api_router.get('/epis/{epi_id}', response_model=EPIResponse)
async def get_epi(epi_id: str, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho', 'almoxarifado'))):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = {"_id": ObjectId(epi_id)}
    empresa_filter = get_empresa_filter(current_user)
    if empresa_filter:
        query.update(empresa_filter)
    
    epi = await db.epis.find_one(query)
    if not epi:
        raise HTTPException(status_code=404, detail='EPI não encontrado')
    stock_status, validity_status = calculate_epi_status(epi)
    resp = doc_to_response(epi)
    resp['stock_status'] = stock_status
    resp['validity_status'] = validity_status
    return EPIResponse(**resp)

@api_router.patch('/epis/{epi_id}', response_model=EPIResponse)
async def update_epi(epi_id: str, epi_data: EPIUpdate, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = {"_id": ObjectId(epi_id)}
    empresa_filter = get_empresa_filter(current_user)
    if empresa_filter:
        query.update(empresa_filter)
    
    update_data = {k: v for k, v in epi_data.model_dump(exclude_unset=True).items()}
    update_data['updated_at'] = datetime.now(timezone.utc)
    
    # Converter enum para string se necessário
    if update_data.get('replacement_period'):
        update_data['replacement_period'] = str(update_data['replacement_period'].value) if hasattr(update_data['replacement_period'], 'value') else str(update_data['replacement_period'])
    
    result = await db.epis.find_one_and_update(query, {"$set": update_data}, return_document=True)
    if not result:
        raise HTTPException(status_code=404, detail='EPI não encontrado')
    stock_status, validity_status = calculate_epi_status(result)
    resp = doc_to_response(result)
    resp['stock_status'] = stock_status
    resp['validity_status'] = validity_status
    return EPIResponse(**resp)

@api_router.delete('/epis/{epi_id}')
async def delete_epi(epi_id: str, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = {"_id": ObjectId(epi_id)}
    empresa_filter = get_empresa_filter(current_user)
    if empresa_filter:
        query.update(empresa_filter)
    
    result = await db.epis.delete_one(query)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='EPI não encontrado')
    return {'message': 'EPI excluído'}

# ===================== KITS =====================

@api_router.get('/kits', response_model=List[KitResponse])
async def get_kits(current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho', 'almoxarifado'))):
    db = await get_db()
    # MULTI-TENANT: Filtrar por empresa
    query = get_empresa_filter(current_user)
    kits = await db.kits.find(query).to_list(1000)
    return [KitResponse(**doc_to_response(k)) for k in kits]

@api_router.post('/kits', response_model=KitResponse, status_code=status.HTTP_201_CREATED)
async def create_kit(kit_data: KitCreate, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    
    # MULTI-TENANT: Associar à empresa do usuário
    empresa_id = current_user.get('empresa_id')
    
    # Buscar detalhes dos EPIs para armazenar nome e descrição
    items_with_details = []
    for item in kit_data.items:
        if item.epi_id:
            epi = await db.epis.find_one({"_id": ObjectId(item.epi_id)})
            if epi:
                items_with_details.append({
                    "epi_id": item.epi_id,
                    "name": epi['name'],
                    "type_category": epi.get('type_category', ''),
                    "ca_number": epi.get('ca_number', ''),
                    "nbr_number": epi.get('nbr_number', ''),
                    "size": epi.get('size', ''),
                    "quantity": item.quantity
                })
    
    new_kit = {
        "name": kit_data.name,
        "description": kit_data.description,
        "sector": kit_data.sector,
        "is_mandatory": kit_data.is_mandatory,
        "items": items_with_details,
        "empresa_id": empresa_id,  # MULTI-TENANT
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    result = await db.kits.insert_one(new_kit)
    new_kit['_id'] = result.inserted_id
    return KitResponse(**doc_to_response(new_kit))

@api_router.get('/kits/{kit_id}', response_model=KitResponse)
async def get_kit(kit_id: str, current_user: dict = Depends(get_current_user)):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = {"_id": ObjectId(kit_id)}
    empresa_filter = get_empresa_filter(current_user)
    if empresa_filter:
        query.update(empresa_filter)
    
    kit = await db.kits.find_one(query)
    if not kit:
        raise HTTPException(status_code=404, detail='Kit não encontrado')
    return KitResponse(**doc_to_response(kit))

@api_router.patch('/kits/{kit_id}', response_model=KitResponse)
async def update_kit(kit_id: str, kit_data: KitUpdate, current_user: dict = Depends(require_role('admin', 'gestor', 'seguranca_trabalho'))):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = {"_id": ObjectId(kit_id)}
    empresa_filter = get_empresa_filter(current_user)
    if empresa_filter:
        query.update(empresa_filter)
    
    update_data = {}
    
    if kit_data.name is not None:
        update_data['name'] = kit_data.name
    if kit_data.description is not None:
        update_data['description'] = kit_data.description
    if kit_data.sector is not None:
        update_data['sector'] = kit_data.sector
    if kit_data.is_mandatory is not None:
        update_data['is_mandatory'] = kit_data.is_mandatory
    
    if kit_data.items is not None:
        items_with_details = []
        for item in kit_data.items:
            if item.epi_id:
                epi = await db.epis.find_one({"_id": ObjectId(item.epi_id)})
                if epi:
                    items_with_details.append({
                        "epi_id": item.epi_id,
                        "name": epi['name'],
                        "type_category": epi.get('type_category', ''),
                        "ca_number": epi.get('ca_number', ''),
                        "nbr_number": epi.get('nbr_number', ''),
                        "size": epi.get('size', ''),
                        "quantity": item.quantity
                    })
        update_data['items'] = items_with_details
    
    update_data['updated_at'] = datetime.now(timezone.utc)
    
    result = await db.kits.find_one_and_update(
        query, {"$set": update_data}, return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail='Kit não encontrado')
    return KitResponse(**doc_to_response(result))

@api_router.delete('/kits/{kit_id}')
async def delete_kit(kit_id: str, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = {"_id": ObjectId(kit_id)}
    empresa_filter = get_empresa_filter(current_user)
    if empresa_filter:
        query.update(empresa_filter)
    
    result = await db.kits.delete_one(query)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Kit não encontrado')
    return {'message': 'Kit excluído'}

# ===================== DELIVERIES =====================

@api_router.post('/deliveries', response_model=DeliveryResponse, status_code=status.HTTP_201_CREATED)
async def create_delivery(delivery_data: DeliveryCreate, current_user: dict = Depends(get_current_user)):
    if not can_deliver_epi(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissão para realizar entregas')
    
    db = await get_db()
    
    # MULTI-TENANT: Buscar colaborador filtrando por empresa
    employee_query = {"_id": ObjectId(delivery_data.employee_id)}
    empresa_id = current_user.get('empresa_id')
    if empresa_id and current_user.get('role') != 'super_admin':
        employee_query["empresa_id"] = empresa_id
    
    employee = await db.employees.find_one(employee_query)
    if not employee:
        raise HTTPException(status_code=404, detail='Colaborador não encontrado')
    
    # Verificar se colaborador tem foto cadastrada
    if not employee.get('photo_path'):
        raise HTTPException(status_code=400, detail='Colaborador não possui foto cadastrada. Procure o RH para cadastrar.')
    
    items_list = []
    for item in delivery_data.items:
        item_dict = item.model_dump()
        
        if item.epi_id:
            # MULTI-TENANT: Buscar EPI filtrando por empresa
            epi_query = {"_id": ObjectId(item.epi_id)}
            if empresa_id and current_user.get('role') != 'super_admin':
                epi_query["empresa_id"] = empresa_id
            
            epi = await db.epis.find_one(epi_query)
            if epi:
                item_dict['epi_name'] = epi['name']
                item_dict['ca_number'] = epi.get('ca_number', '')
                stock_change = -item.quantity if not delivery_data.is_return else item.quantity
                await db.epis.update_one({"_id": ObjectId(item.epi_id)}, {"$inc": {"current_stock": stock_change}})
                
                movement = {
                    "movement_type": "return" if delivery_data.is_return else "delivery",
                    "epi_id": item.epi_id,
                    "quantity": item.quantity if delivery_data.is_return else -item.quantity,
                    "employee_id": delivery_data.employee_id,
                    "empresa_id": empresa_id,  # MULTI-TENANT
                    "created_by": current_user['id'],
                    "created_at": datetime.now(timezone.utc)
                }
                await db.stock_movements.insert_one(movement)
        
        if item.kit_id:
            # MULTI-TENANT: Buscar Kit filtrando por empresa
            kit_query = {"_id": ObjectId(item.kit_id)}
            if empresa_id and current_user.get('role') != 'super_admin':
                kit_query["empresa_id"] = empresa_id
            
            kit = await db.kits.find_one(kit_query)
            if kit:
                item_dict['kit_name'] = kit['name']
                # Processar itens do kit
                for kit_item in kit.get('items', []):
                    if kit_item.get('epi_id'):
                        stock_change = -kit_item['quantity'] if not delivery_data.is_return else kit_item['quantity']
                        await db.epis.update_one({"_id": ObjectId(kit_item['epi_id'])}, {"$inc": {"current_stock": stock_change}})
        
        items_list.append(item_dict)
    
    new_delivery = {
        "employee_id": delivery_data.employee_id,
        "employee_name": employee['full_name'],
        "delivery_type": delivery_data.delivery_type,
        "is_return": delivery_data.is_return,
        "facial_match_score": delivery_data.facial_match_score,
        "facial_photo_path": delivery_data.facial_photo_path,
        "notes": delivery_data.notes,
        "items": items_list,
        "delivered_by": current_user['id'],
        "delivered_by_name": current_user['username'],
        "empresa_id": current_user.get('empresa_id'),  # MULTI-TENANT
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.deliveries.insert_one(new_delivery)
    new_delivery['_id'] = result.inserted_id
    return DeliveryResponse(**doc_to_response(new_delivery))

@api_router.post('/deliveries/save-photo')
async def save_delivery_photo(
    employee_id: str = Form(...),
    photo_data: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Salva a foto de confirmação da entrega"""
    if not can_deliver_epi(current_user['role']):
        raise HTTPException(status_code=403, detail='Sem permissão')
    
    try:
        # Decodificar base64
        if ',' in photo_data:
            photo_data = photo_data.split(',')[1]
        
        photo_bytes = base64.b64decode(photo_data)
        
        file_name = f'delivery_{employee_id}_{datetime.now(timezone.utc).timestamp()}.jpg'
        file_path = UPLOAD_DIR / 'deliveries' / file_name
        file_path.parent.mkdir(exist_ok=True, parents=True)
        
        with open(file_path, 'wb') as f:
            f.write(photo_bytes)
        
        return {'photo_path': f'/uploads/deliveries/{file_name}'}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'Erro ao salvar foto: {str(e)}')

@api_router.get('/deliveries', response_model=List[DeliveryResponse])
async def get_deliveries(
    current_user: dict = Depends(get_current_user), 
    employee_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = get_empresa_filter(current_user)
    
    if employee_id:
        query['employee_id'] = employee_id
    
    if start_date:
        query['created_at'] = query.get('created_at', {})
        query['created_at']['$gte'] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
    
    if end_date:
        query['created_at'] = query.get('created_at', {})
        query['created_at']['$lte'] = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    
    deliveries = await db.deliveries.find(query).sort("created_at", -1).to_list(1000)
    return [DeliveryResponse(**doc_to_response(d)) for d in deliveries]

# ===================== STOCK =====================

@api_router.get('/stock/alerts')
async def get_stock_alerts(current_user: dict = Depends(get_current_user)):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    empresa_filter = get_empresa_filter(current_user)
    
    # EPIs com estoque baixo
    low_stock_query = {"$expr": {"$lte": ["$current_stock", "$min_stock"]}}
    if empresa_filter:
        low_stock_query.update(empresa_filter)
    low_stock = await db.epis.find(low_stock_query).to_list(100)
    
    # EPIs com validade próxima (30 dias)
    expiry_date = datetime.now(timezone.utc) + timedelta(days=30)
    expiring_query = {
        "$or": [
            {"validity_date": {"$ne": None, "$lte": expiry_date}},
            {"ca_validity": {"$ne": None, "$lte": expiry_date}}
        ]
    }
    if empresa_filter:
        expiring_query.update(empresa_filter)
    expiring_soon = await db.epis.find(expiring_query).to_list(100)
    
    return {
        'low_stock': [{'id': str(e['_id']), 'name': e['name'], 'current_stock': e['current_stock'], 'min_stock': e['min_stock']} for e in low_stock],
        'expiring_soon': [{'id': str(e['_id']), 'name': e['name'], 'validity_date': e.get('validity_date') or e.get('ca_validity')} for e in expiring_soon]
    }

@api_router.get('/stock/movements')
async def get_stock_movements(current_user: dict = Depends(get_current_user), epi_id: Optional[str] = None):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = get_empresa_filter(current_user)
    
    if epi_id:
        query['epi_id'] = epi_id
    movements = await db.stock_movements.find(query).sort("created_at", -1).to_list(500)
    return [doc_to_response(m) for m in movements]

# ===================== LICENSE =====================

@api_router.get('/license', response_model=LicenseResponse)
async def get_license(current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    license_doc = await db.panel_license.find_one({})
    if not license_doc:
        raise HTTPException(status_code=404, detail='Licença não encontrada')
    
    now = datetime.now(timezone.utc)
    expires_at = license_doc['expires_at']
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    days_remaining = max(0, (expires_at - now).days)
    
    return LicenseResponse(
        id=str(license_doc['_id']),
        expires_at=license_doc['expires_at'],
        is_blocked=license_doc.get('is_blocked', False),
        days_remaining=days_remaining
    )

@api_router.post('/license/add-days')
async def add_license_days(request: LicenseAddDaysRequest, current_user: dict = Depends(require_role('admin'))):
    db = await get_db()
    license_doc = await db.panel_license.find_one({})
    if not license_doc:
        raise HTTPException(status_code=404, detail='Licença não encontrada')
    
    new_expires = license_doc['expires_at'] + timedelta(days=request.days)
    await db.panel_license.update_one({"_id": license_doc['_id']}, {"$set": {"expires_at": new_expires}})
    
    history = {
        "license_id": str(license_doc['_id']),
        "user_id": current_user['id'],
        "days_added": request.days,
        "reason": request.reason,
        "created_at": datetime.now(timezone.utc)
    }
    await db.license_history.insert_one(history)
    
    return {'message': f'{request.days} dias adicionados com sucesso'}

# ===================== ALERTAS DE EPI OBRIGATÓRIO E PERIODICIDADE =====================

def get_replacement_days(epi):
    """Retorna o número de dias para troca baseado na periodicidade"""
    period = epi.get('replacement_period')
    if period == 'weekly':
        return 7
    elif period == 'biweekly':
        return 14
    elif period == 'monthly':
        return 30
    elif period == 'custom':
        return epi.get('replacement_days', 30)
    return None

@api_router.get('/alerts/pending-epis')
async def get_pending_epi_alerts(current_user: dict = Depends(get_current_user)):
    """Retorna alertas de EPIs obrigatórios não entregues aos colaboradores"""
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    empresa_filter = get_empresa_filter(current_user)
    
    alerts = []
    
    # Buscar colaboradores ativos DA EMPRESA
    emp_query = {"status": "active"}
    if empresa_filter:
        emp_query.update(empresa_filter)
    employees = await db.employees.find(emp_query).to_list(5000)
    
    # Buscar kits obrigatórios DA EMPRESA
    kit_query = {"is_mandatory": {"$ne": False}}
    if empresa_filter:
        kit_query.update(empresa_filter)
    kits = await db.kits.find(kit_query).to_list(100)
    kits_by_sector = {kit.get('sector', '').lower(): kit for kit in kits if kit.get('sector')}
    
    # Buscar entregas DA EMPRESA
    delivery_query = {"is_return": False}
    if empresa_filter:
        delivery_query.update(empresa_filter)
    deliveries = await db.deliveries.find(delivery_query).to_list(10000)
    
    # Agrupar entregas por colaborador
    deliveries_by_employee = {}
    for d in deliveries:
        emp_id = d.get('employee_id')
        if emp_id not in deliveries_by_employee:
            deliveries_by_employee[emp_id] = []
        deliveries_by_employee[emp_id].append(d)
    
    for employee in employees:
        emp_id = str(employee['_id'])
        department = (employee.get('department') or '').lower()
        
        # Verificar se há kit obrigatório para o setor
        kit = kits_by_sector.get(department)
        if not kit:
            continue
        
        # Buscar EPIs já entregues ao colaborador
        emp_deliveries = deliveries_by_employee.get(emp_id, [])
        delivered_epi_ids = set()
        for d in emp_deliveries:
            for item in d.get('items', []):
                if item.get('epi_id'):
                    delivered_epi_ids.add(item.get('epi_id'))
        
        # Verificar EPIs faltantes do kit
        missing_epis = []
        for kit_item in kit.get('items', []):
            epi_id = kit_item.get('epi_id')
            if epi_id and epi_id not in delivered_epi_ids:
                missing_epis.append({
                    'epi_id': epi_id,
                    'name': kit_item.get('name', 'EPI'),
                    'ca_number': kit_item.get('ca_number', ''),
                    'quantity_required': kit_item.get('quantity', 1)
                })
        
        if missing_epis:
            alerts.append({
                'employee_id': emp_id,
                'employee_name': employee.get('full_name', ''),
                'department': employee.get('department', ''),
                'kit_name': kit.get('name', ''),
                'kit_id': str(kit['_id']),
                'missing_epis': missing_epis,
                'alert_type': 'missing_mandatory_epi',
                'message': f"{employee.get('full_name')} ({employee.get('department', '')}) não possui {len(missing_epis)} EPI(s) obrigatório(s) do kit"
            })
    
    return alerts

@api_router.get('/alerts/replacement-due')
async def get_replacement_due_alerts(current_user: dict = Depends(get_current_user)):
    """Retorna alertas de EPIs com troca periódica vencida"""
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    empresa_filter = get_empresa_filter(current_user)
    
    alerts = []
    now = datetime.now(timezone.utc)
    
    # Buscar EPIs com periodicidade de troca definida DA EMPRESA
    epi_query = {"replacement_period": {"$ne": None}}
    if empresa_filter:
        epi_query.update(empresa_filter)
    epis_with_period = await db.epis.find(epi_query).to_list(500)
    
    epi_periods = {str(e['_id']): e for e in epis_with_period}
    
    if not epi_periods:
        return alerts
    
    # Buscar colaboradores ativos DA EMPRESA
    emp_query = {"status": "active"}
    if empresa_filter:
        emp_query.update(empresa_filter)
    employees = await db.employees.find(emp_query).to_list(5000)
    emp_map = {str(e['_id']): e for e in employees}
    
    # Buscar entregas mais recentes de EPIs com periodicidade
    epi_ids = list(epi_periods.keys())
    
    # Usar agregação para encontrar última entrega de cada EPI por colaborador
    match_query = {"is_return": False}
    if empresa_filter:
        match_query.update(empresa_filter)
    
    pipeline = [
        {"$match": match_query},
        {"$unwind": "$items"},
        {"$match": {"items.epi_id": {"$in": epi_ids}}},
        {"$sort": {"created_at": -1}},
        {"$group": {
            "_id": {"employee_id": "$employee_id", "epi_id": "$items.epi_id"},
            "last_delivery": {"$first": "$created_at"},
            "employee_name": {"$first": "$employee_name"}
        }}
    ]
    
    results = await db.deliveries.aggregate(pipeline).to_list(10000)
    
    for result in results:
        emp_id = result['_id']['employee_id']
        epi_id = result['_id']['epi_id']
        last_delivery = result['last_delivery']
        
        if epi_id not in epi_periods:
            continue
        
        epi = epi_periods[epi_id]
        replacement_days = get_replacement_days(epi)
        
        if replacement_days is None:
            continue
        
        # Calcular data de vencimento da troca
        if last_delivery.tzinfo is None:
            last_delivery = last_delivery.replace(tzinfo=timezone.utc)
        
        due_date = last_delivery + timedelta(days=replacement_days)
        
        if now > due_date:
            days_overdue = (now - due_date).days
            employee = emp_map.get(emp_id, {})
            
            alerts.append({
                'employee_id': emp_id,
                'employee_name': result.get('employee_name', employee.get('full_name', '')),
                'epi_id': epi_id,
                'epi_name': epi.get('name', ''),
                'ca_number': epi.get('ca_number', ''),
                'nbr_number': epi.get('nbr_number', ''),
                'last_delivery_date': last_delivery.isoformat(),
                'replacement_due_date': due_date.isoformat(),
                'days_overdue': days_overdue,
                'replacement_period': epi.get('replacement_period', ''),
                'alert_type': 'replacement_due',
                'message': f"{result.get('employee_name', 'Colaborador')} não realizou retirada de {epi.get('name')} - {days_overdue} dias de atraso"
            })
    
    return alerts

@api_router.get('/alerts/all')
async def get_all_alerts(current_user: dict = Depends(get_current_user)):
    """Retorna todos os alertas consolidados"""
    pending_epis = await get_pending_epi_alerts(current_user)
    replacement_due = await get_replacement_due_alerts(current_user)
    
    return {
        'pending_epis': pending_epis,
        'replacement_due': replacement_due,
        'total_pending_epis': len(pending_epis),
        'total_replacement_due': len(replacement_due),
        'total_alerts': len(pending_epis) + len(replacement_due)
    }

@api_router.get('/alerts/employee/{employee_id}')
async def get_employee_alerts(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Retorna alertas específicos de um colaborador"""
    db = await get_db()
    
    employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    
    alerts = {
        'pending_epis': [],
        'replacement_due': []
    }
    
    department = (employee.get('department') or '').lower()
    
    # Buscar kit obrigatório do setor
    kit = await db.kits.find_one({
        "sector": {"$regex": f"^{department}$", "$options": "i"},
        "is_mandatory": {"$ne": False}
    })
    
    if kit:
        # Buscar EPIs já entregues
        deliveries = await db.deliveries.find({
            "employee_id": employee_id,
            "is_return": False
        }).to_list(1000)
        
        delivered_epi_ids = set()
        for d in deliveries:
            for item in d.get('items', []):
                if item.get('epi_id'):
                    delivered_epi_ids.add(item.get('epi_id'))
        
        # Verificar EPIs faltantes
        for kit_item in kit.get('items', []):
            epi_id = kit_item.get('epi_id')
            if epi_id and epi_id not in delivered_epi_ids:
                alerts['pending_epis'].append({
                    'epi_id': epi_id,
                    'name': kit_item.get('name', 'EPI'),
                    'ca_number': kit_item.get('ca_number', ''),
                    'quantity_required': kit_item.get('quantity', 1),
                    'message': f"EPI obrigatório não entregue: {kit_item.get('name', 'EPI')}"
                })
    
    # Verificar EPIs com troca vencida
    epi_ids_with_period = []
    epis_with_period = await db.epis.find({"replacement_period": {"$ne": None}}).to_list(500)
    epi_periods = {str(e['_id']): e for e in epis_with_period}
    
    if epi_periods:
        now = datetime.now(timezone.utc)
        
        # Buscar última entrega de cada EPI com periodicidade
        pipeline = [
            {"$match": {"employee_id": employee_id, "is_return": False}},
            {"$unwind": "$items"},
            {"$match": {"items.epi_id": {"$in": list(epi_periods.keys())}}},
            {"$sort": {"created_at": -1}},
            {"$group": {
                "_id": "$items.epi_id",
                "last_delivery": {"$first": "$created_at"}
            }}
        ]
        
        results = await db.deliveries.aggregate(pipeline).to_list(500)
        
        for result in results:
            epi_id = result['_id']
            last_delivery = result['last_delivery']
            
            if epi_id not in epi_periods:
                continue
            
            epi = epi_periods[epi_id]
            replacement_days = get_replacement_days(epi)
            
            if replacement_days is None:
                continue
            
            if last_delivery.tzinfo is None:
                last_delivery = last_delivery.replace(tzinfo=timezone.utc)
            
            due_date = last_delivery + timedelta(days=replacement_days)
            
            if now > due_date:
                days_overdue = (now - due_date).days
                alerts['replacement_due'].append({
                    'epi_id': epi_id,
                    'epi_name': epi.get('name', ''),
                    'last_delivery_date': last_delivery.isoformat(),
                    'replacement_due_date': due_date.isoformat(),
                    'days_overdue': days_overdue,
                    'message': f"Troca vencida há {days_overdue} dias: {epi.get('name', '')}"
                })
    
    alerts['total_alerts'] = len(alerts['pending_epis']) + len(alerts['replacement_due'])
    alerts['kit_name'] = kit.get('name') if kit else None
    alerts['kit_id'] = str(kit['_id']) if kit else None
    
    return alerts

# ===================== SETOR-KIT VINCULAÇÃO =====================

@api_router.get('/kits/by-sector/{sector_name}')
async def get_kit_by_sector(sector_name: str, current_user: dict = Depends(get_current_user)):
    """Retorna o kit obrigatório vinculado a um setor"""
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    query = {
        "sector": {"$regex": f"^{sector_name}$", "$options": "i"},
        "is_mandatory": {"$ne": False}
    }
    empresa_filter = get_empresa_filter(current_user)
    if empresa_filter:
        query.update(empresa_filter)
    
    kit = await db.kits.find_one(query)
    
    if not kit:
        return None
    
    return KitResponse(**doc_to_response(kit))

@api_router.get('/sectors/list')
async def get_sectors_list(current_user: dict = Depends(get_current_user)):
    """Retorna lista de setores com seus kits vinculados"""
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    empresa_filter = get_empresa_filter(current_user)
    
    # Buscar setores dos colaboradores DA EMPRESA
    match_stage = {"department": {"$ne": None, "$ne": ""}}
    if empresa_filter:
        match_stage.update(empresa_filter)
    
    pipeline = [
        {"$match": match_stage},
        {"$group": {"_id": "$department"}},
        {"$sort": {"_id": 1}}
    ]
    
    sectors_result = await db.employees.aggregate(pipeline).to_list(100)
    sectors = [s['_id'] for s in sectors_result if s['_id']]
    
    # Buscar kits DA EMPRESA
    kit_query = {}
    if empresa_filter:
        kit_query.update(empresa_filter)
    kits = await db.kits.find(kit_query).to_list(100)
    kits_by_sector = {(k.get('sector') or '').lower(): k for k in kits}
    
    result = []
    for sector in sectors:
        kit = kits_by_sector.get(sector.lower())
        result.append({
            'sector_name': sector,
            'has_kit': kit is not None,
            'kit_id': str(kit['_id']) if kit else None,
            'kit_name': kit.get('name') if kit else None,
            'kit_items_count': len(kit.get('items', [])) if kit else 0
        })
    
    return result

# ===================== DASHBOARD =====================

@api_router.get('/dashboard/stats')
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    empresa_filter = get_empresa_filter(current_user)
    
    # Colaboradores ativos DA EMPRESA
    emp_query = {"status": "active"}
    if empresa_filter:
        emp_query.update(empresa_filter)
    active_employees = await db.employees.count_documents(emp_query)
    
    # EPIs DA EMPRESA
    epi_query = {}
    if empresa_filter:
        epi_query.update(empresa_filter)
    total_epis = await db.epis.count_documents(epi_query)
    
    # EPIs com estoque baixo DA EMPRESA
    low_stock_query = {"$expr": {"$lte": ["$current_stock", "$min_stock"]}}
    if empresa_filter:
        low_stock_query.update(empresa_filter)
    low_stock_count = await db.epis.count_documents(low_stock_query)
    
    # Entregas recentes DA EMPRESA
    thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
    delivery_query = {
        "is_return": False,
        "created_at": {"$gte": thirty_days_ago}
    }
    if empresa_filter:
        delivery_query.update(empresa_filter)
    recent_deliveries = await db.deliveries.count_documents(delivery_query)
    
    # EPIs com validade próxima DA EMPRESA
    expiry_date = datetime.now(timezone.utc) + timedelta(days=30)
    expiring_query = {
        "$or": [
            {"validity_date": {"$ne": None, "$lte": expiry_date}},
            {"ca_validity": {"$ne": None, "$lte": expiry_date}}
        ]
    }
    if empresa_filter:
        expiring_query.update(empresa_filter)
    expiring_count = await db.epis.count_documents(expiring_query)
    
    # NOVO: Contagem de alertas de EPIs obrigatórios pendentes
    pending_epi_alerts = 0
    replacement_due_alerts = 0
    try:
        # Contar alertas de EPIs pendentes (simplificado para performance)
        kit_query = {
            "sector": {"$ne": None, "$ne": ""},
            "is_mandatory": {"$ne": False}
        }
        if empresa_filter:
            kit_query.update(empresa_filter)
        kits_with_sectors = await db.kits.count_documents(kit_query)
        
        # Contar EPIs com periodicidade vencida DA EMPRESA
        period_query = {"replacement_period": {"$ne": None}}
        if empresa_filter:
            period_query.update(empresa_filter)
        epis_with_period = await db.epis.count_documents(period_query)
        
        # Se houver kits ou EPIs com periodicidade, buscar alertas completos
        if kits_with_sectors > 0 or epis_with_period > 0:
            all_alerts = await get_all_alerts(current_user)
            pending_epi_alerts = all_alerts.get('total_pending_epis', 0)
            replacement_due_alerts = all_alerts.get('total_replacement_due', 0)
    except Exception as e:
        logger.error(f"Erro ao buscar alertas: {e}")
    
    return {
        'active_employees': active_employees,
        'total_epis': total_epis,
        'low_stock_count': low_stock_count,
        'recent_deliveries': recent_deliveries,
        'expiring_epis': expiring_count,
        'pending_epi_alerts': pending_epi_alerts,
        'replacement_due_alerts': replacement_due_alerts,
        'total_alerts': pending_epi_alerts + replacement_due_alerts
    }

# ===================== AUTENTICAÇÃO DE FICHA EPI =====================

import hashlib
import secrets

def generate_auth_code(employee_id: str, timestamp: datetime) -> str:
    """Gera código de autenticação único para ficha de EPI"""
    # Criar hash baseado em employee_id, timestamp e secret
    secret_key = os.environ.get('SECRET_KEY', 'gestorepi-secret-key-2026')
    data = f"{employee_id}:{timestamp.isoformat()}:{secret_key}:{secrets.token_hex(4)}"
    hash_obj = hashlib.sha256(data.encode())
    hash_hex = hash_obj.hexdigest()[:10].upper()
    return f"AUT-EPI-{hash_hex}"

@api_router.post('/ficha-auth/generate', response_model=FichaAuthenticationResponse)
async def generate_ficha_authentication(
    data: FichaAuthenticationCreate,
    current_user: dict = Depends(get_current_user)
):
    """Gera código de autenticação para ficha de EPI com validação biométrica"""
    db = await get_db()
    
    # Buscar colaborador
    employee = await db.employees.find_one({"_id": ObjectId(data.employee_id)})
    if not employee:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    
    # Buscar entregas do colaborador
    delivery_query = {"employee_id": data.employee_id}
    if not data.include_all_history and data.delivery_ids:
        delivery_query["_id"] = {"$in": [ObjectId(d) for d in data.delivery_ids]}
    
    deliveries = await db.deliveries.find(delivery_query).to_list(1000)
    delivery_ids = [str(d["_id"]) for d in deliveries]
    
    # Verificar se há validação biométrica nas entregas
    biometric_validated = False
    max_biometric_score = 0.0
    
    for delivery in deliveries:
        if delivery.get("facial_match_score"):
            biometric_validated = True
            score = delivery.get("facial_match_score", 0)
            if score > max_biometric_score:
                max_biometric_score = score
    
    # Gerar código de autenticação
    now = datetime.now(timezone.utc)
    auth_code = generate_auth_code(data.employee_id, now)
    
    # Dados para QR Code (JSON que pode ser escaneado para verificação)
    qr_data = {
        "type": "EPI_FICHA_AUTH",
        "code": auth_code,
        "employee_id": data.employee_id,
        "employee_name": employee.get("full_name"),
        "date": now.isoformat(),
        "biometric": biometric_validated,
        "verify_url": f"/api/ficha-auth/verify/{auth_code}"
    }
    
    # Salvar no banco
    auth_record = {
        "auth_code": auth_code,
        "employee_id": data.employee_id,
        "employee_name": employee.get("full_name"),
        "delivery_ids": delivery_ids,
        "biometric_validated": biometric_validated,
        "biometric_score": max_biometric_score if biometric_validated else None,
        "created_at": now,
        "created_by": current_user.get("username"),
        "created_by_id": current_user.get("id"),
        "qr_code_data": str(qr_data),
        "status": "active"
    }
    
    result = await db.ficha_authentications.insert_one(auth_record)
    
    return FichaAuthenticationResponse(
        id=str(result.inserted_id),
        auth_code=auth_code,
        employee_id=data.employee_id,
        employee_name=employee.get("full_name"),
        delivery_ids=delivery_ids,
        biometric_validated=biometric_validated,
        biometric_score=max_biometric_score if biometric_validated else None,
        created_at=now,
        created_by=current_user.get("username"),
        qr_code_data=str(qr_data)
    )

@api_router.get('/ficha-auth/verify/{auth_code}', response_model=FichaAuthenticationVerify)
async def verify_ficha_authentication(auth_code: str):
    """Verifica autenticidade de uma ficha de EPI pelo código"""
    db = await get_db()
    
    # Buscar registro de autenticação
    auth_record = await db.ficha_authentications.find_one({"auth_code": auth_code})
    
    if not auth_record:
        return FichaAuthenticationVerify(
            valid=False,
            auth_code=auth_code,
            message="Código de autenticação não encontrado no sistema"
        )
    
    if auth_record.get("status") == "revoked":
        return FichaAuthenticationVerify(
            valid=False,
            auth_code=auth_code,
            employee_name=auth_record.get("employee_name"),
            validation_date=auth_record.get("created_at"),
            message="Este código foi revogado e não é mais válido"
        )
    
    return FichaAuthenticationVerify(
        valid=True,
        auth_code=auth_code,
        employee_name=auth_record.get("employee_name"),
        validation_date=auth_record.get("created_at"),
        biometric_validated=auth_record.get("biometric_validated"),
        message="Documento autêntico - validado pelo sistema GestorEPI"
    )

@api_router.get('/ficha-auth/employee/{employee_id}')
async def get_employee_authentications(
    employee_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Lista todas as autenticações de ficha de um colaborador"""
    db = await get_db()
    
    records = await db.ficha_authentications.find(
        {"employee_id": employee_id}
    ).sort("created_at", -1).to_list(100)
    
    return [{
        "id": str(r["_id"]),
        "auth_code": r["auth_code"],
        "created_at": r["created_at"],
        "biometric_validated": r.get("biometric_validated"),
        "status": r.get("status", "active")
    } for r in records]

# ===================== FASE 3: PAINEL MASTER COMPLETO - RELATÓRIOS =====================

@api_router.get('/master/dashboard')
async def get_master_dashboard(current_user: dict = Depends(require_super_admin())):
    """Dashboard geral do Painel Master com métricas de todas empresas"""
    db = await get_db()
    
    # Total de empresas
    total_empresas = await db.empresas.count_documents({})
    empresas_ativas = await db.empresas.count_documents({"status": "ativo"})
    empresas_bloqueadas = await db.empresas.count_documents({"status": "bloqueado"})
    
    # Total de colaboradores no sistema
    total_colaboradores = await db.employees.count_documents({})
    total_entregas = await db.deliveries.count_documents({})
    
    # Empresas por plano
    pipeline_planos = [
        {"$group": {"_id": "$plano", "count": {"$sum": 1}}}
    ]
    planos_result = await db.empresas.aggregate(pipeline_planos).to_list(10)
    empresas_por_plano = {p['_id']: p['count'] for p in planos_result}
    
    # Buscar todas as empresas para calcular alertas de limite
    empresas = await db.empresas.find({"status": "ativo"}).to_list(1000)
    
    empresas_warning = 0
    empresas_critical = 0
    total_uso = 0
    
    for emp in empresas:
        emp_id = str(emp['_id'])
        colab_count = await db.employees.count_documents({"empresa_id": emp_id})
        limite = emp.get('limite_colaboradores', 50)
        uso = (colab_count / limite * 100) if limite > 0 else 0
        total_uso += uso
        
        if uso >= 90:
            empresas_critical += 1
        elif uso >= 80:
            empresas_warning += 1
    
    media_uso = total_uso / len(empresas) if empresas else 0
    
    # Novos colaboradores e empresas no último mês
    mes_atras = datetime.now(timezone.utc) - timedelta(days=30)
    novos_colaboradores = await db.employees.count_documents({"created_at": {"$gte": mes_atras}})
    novas_empresas = await db.empresas.count_documents({"created_at": {"$gte": mes_atras}})
    
    return {
        "total_empresas": total_empresas,
        "empresas_ativas": empresas_ativas,
        "empresas_bloqueadas": empresas_bloqueadas,
        "total_colaboradores_sistema": total_colaboradores,
        "total_entregas_sistema": total_entregas,
        "empresas_por_plano": empresas_por_plano,
        "empresas_limite_warning": empresas_warning,
        "empresas_limite_critical": empresas_critical,
        "media_uso_plano": round(media_uso, 1),
        "novos_colaboradores_mes": novos_colaboradores,
        "novas_empresas_mes": novas_empresas
    }

@api_router.get('/master/alertas-limite')
async def get_alertas_limite(current_user: dict = Depends(require_super_admin())):
    """Lista empresas que estão próximas ou no limite do plano"""
    db = await get_db()
    
    empresas = await db.empresas.find({"status": "ativo"}).to_list(1000)
    alertas = []
    
    for emp in empresas:
        emp_id = str(emp['_id'])
        colab_count = await db.employees.count_documents({"empresa_id": emp_id})
        limite = emp.get('limite_colaboradores', 50)
        uso = (colab_count / limite * 100) if limite > 0 else 0
        
        if uso >= 80:
            if uso >= 100:
                nivel = "blocked"
                msg = f"BLOQUEADO: Limite atingido ({colab_count}/{limite})"
            elif uso >= 90:
                nivel = "critical"
                msg = f"CRÍTICO: {uso:.0f}% do limite ({colab_count}/{limite})"
            else:
                nivel = "warning"
                msg = f"ATENÇÃO: {uso:.0f}% do limite ({colab_count}/{limite})"
            
            alertas.append({
                "empresa_id": emp_id,
                "empresa_nome": emp.get('nome'),
                "colaboradores_atual": colab_count,
                "limite": limite,
                "uso_percentual": round(uso, 1),
                "nivel_alerta": nivel,
                "mensagem": msg,
                "plano": emp.get('plano')
            })
    
    # Ordenar por uso percentual decrescente
    alertas.sort(key=lambda x: x['uso_percentual'], reverse=True)
    
    return {
        "alertas": alertas,
        "total_warning": len([a for a in alertas if a['nivel_alerta'] == 'warning']),
        "total_critical": len([a for a in alertas if a['nivel_alerta'] == 'critical']),
        "total_blocked": len([a for a in alertas if a['nivel_alerta'] == 'blocked'])
    }

@api_router.get('/master/empresas/{empresa_id}/relatorio')
async def get_empresa_relatorio(
    empresa_id: str,
    periodo: int = 30,
    current_user: dict = Depends(require_super_admin())
):
    """Relatório detalhado de uma empresa específica"""
    db = await get_db()
    
    empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    data_inicio = datetime.now(timezone.utc) - timedelta(days=periodo)
    
    # Colaboradores
    total_colab = await db.employees.count_documents({"empresa_id": empresa_id})
    colab_ativos = await db.employees.count_documents({"empresa_id": empresa_id, "status": "active"})
    colab_inativos = total_colab - colab_ativos
    
    # EPIs
    total_epis = await db.epis.count_documents({"empresa_id": empresa_id})
    epis_baixo = await db.epis.count_documents({
        "empresa_id": empresa_id,
        "$expr": {"$lte": ["$current_stock", "$min_stock"]}
    })
    
    data_vencimento = datetime.now(timezone.utc) + timedelta(days=30)
    epis_vencendo = await db.epis.count_documents({
        "empresa_id": empresa_id,
        "$or": [
            {"validity_date": {"$lte": data_vencimento}},
            {"ca_validity": {"$lte": data_vencimento}}
        ]
    })
    
    # Entregas
    total_entregas = await db.deliveries.count_documents({"empresa_id": empresa_id})
    entregas_periodo = await db.deliveries.count_documents({
        "empresa_id": empresa_id,
        "is_return": False,
        "created_at": {"$gte": data_inicio}
    })
    devolucoes_periodo = await db.deliveries.count_documents({
        "empresa_id": empresa_id,
        "is_return": True,
        "created_at": {"$gte": data_inicio}
    })
    
    # Usuários
    total_usuarios = await db.users.count_documents({"empresa_id": empresa_id})
    
    # Kits
    total_kits = await db.kits.count_documents({"empresa_id": empresa_id})
    
    # Entregas por dia (últimos 7 dias)
    pipeline_diario = [
        {
            "$match": {
                "empresa_id": empresa_id,
                "is_return": False,
                "created_at": {"$gte": datetime.now(timezone.utc) - timedelta(days=7)}
            }
        },
        {
            "$group": {
                "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$created_at"}},
                "count": {"$sum": 1}
            }
        },
        {"$sort": {"_id": 1}}
    ]
    entregas_diarias = await db.deliveries.aggregate(pipeline_diario).to_list(7)
    
    # Top EPIs mais entregues
    pipeline_top_epis = [
        {"$match": {"empresa_id": empresa_id, "is_return": False, "created_at": {"$gte": data_inicio}}},
        {"$unwind": "$items"},
        {"$group": {"_id": "$items.epi_name", "total": {"$sum": "$items.quantity"}}},
        {"$sort": {"total": -1}},
        {"$limit": 5}
    ]
    top_epis = await db.deliveries.aggregate(pipeline_top_epis).to_list(5)
    
    # Top colaboradores com mais entregas
    pipeline_top_colab = [
        {"$match": {"empresa_id": empresa_id, "is_return": False, "created_at": {"$gte": data_inicio}}},
        {"$group": {"_id": "$employee_name", "total": {"$sum": 1}}},
        {"$sort": {"total": -1}},
        {"$limit": 5}
    ]
    top_colaboradores = await db.deliveries.aggregate(pipeline_top_colab).to_list(5)
    
    limite = empresa.get('limite_colaboradores', 50)
    uso_percentual = (total_colab / limite * 100) if limite > 0 else 0
    
    return {
        "empresa_id": empresa_id,
        "empresa_nome": empresa.get('nome'),
        "periodo_dias": periodo,
        "total_colaboradores": total_colab,
        "colaboradores_ativos": colab_ativos,
        "colaboradores_inativos": colab_inativos,
        "limite_colaboradores": limite,
        "uso_percentual": round(uso_percentual, 1),
        "total_epis": total_epis,
        "epis_estoque_baixo": epis_baixo,
        "epis_vencendo": epis_vencendo,
        "total_entregas": total_entregas,
        "entregas_periodo": entregas_periodo,
        "devolucoes_periodo": devolucoes_periodo,
        "total_usuarios": total_usuarios,
        "total_kits": total_kits,
        "alertas_pendentes": 0,  # TODO: calcular alertas específicos
        "entregas_por_dia": [{"data": e['_id'], "total": e['count']} for e in entregas_diarias],
        "top_epis_entregues": [{"nome": e['_id'], "total": e['total']} for e in top_epis if e['_id']],
        "top_colaboradores_entregas": [{"nome": c['_id'], "total": c['total']} for c in top_colaboradores if c['_id']],
        "gerado_em": datetime.now(timezone.utc).isoformat()
    }

@api_router.get('/master/empresas/{empresa_id}/export/pdf')
async def export_empresa_relatorio_pdf(
    empresa_id: str,
    periodo: int = 30,
    current_user: dict = Depends(require_super_admin())
):
    """Exporta relatório da empresa em PDF"""
    db = await get_db()
    
    empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    # Buscar dados do relatório
    relatorio = await get_empresa_relatorio(empresa_id, periodo, current_user)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=40)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, spaceAfter=10)
    
    elements = []
    
    # Cabeçalho
    elements.append(Paragraph(f"Relatório da Empresa", title_style))
    elements.append(Paragraph(f"{empresa.get('nome')}", subtitle_style))
    elements.append(Paragraph(f"Período: últimos {periodo} dias | Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Informações da Empresa
    elements.append(Paragraph("Informações da Empresa", subtitle_style))
    info_data = [
        ['CNPJ:', empresa.get('cnpj', '-')],
        ['Plano:', empresa.get('plano', '-')],
        ['Status:', empresa.get('status', '-')],
        ['Responsável:', empresa.get('responsavel', '-')],
        ['E-mail:', empresa.get('email', '-')],
        ['Telefone:', empresa.get('telefone', '-')],
    ]
    info_table = Table(info_data, colWidths=[120, 350])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Métricas
    elements.append(Paragraph("Métricas de Uso", subtitle_style))
    metrics_data = [
        ['Métrica', 'Valor'],
        ['Colaboradores', f"{relatorio['colaboradores_ativos']} ativos / {relatorio['total_colaboradores']} total"],
        ['Uso do Plano', f"{relatorio['uso_percentual']}% ({relatorio['total_colaboradores']}/{relatorio['limite_colaboradores']})"],
        ['EPIs Cadastrados', str(relatorio['total_epis'])],
        ['EPIs Estoque Baixo', str(relatorio['epis_estoque_baixo'])],
        ['EPIs Vencendo', str(relatorio['epis_vencendo'])],
        ['Entregas no Período', str(relatorio['entregas_periodo'])],
        ['Devoluções no Período', str(relatorio['devolucoes_periodo'])],
        ['Usuários', str(relatorio['total_usuarios'])],
        ['Kits', str(relatorio['total_kits'])],
    ]
    
    metrics_table = Table(metrics_data, colWidths=[200, 270])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.4, 0.2, 0.6)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
    ]))
    elements.append(metrics_table)
    elements.append(Spacer(1, 20))
    
    # Top EPIs
    if relatorio['top_epis_entregues']:
        elements.append(Paragraph("Top 5 EPIs Mais Entregues", subtitle_style))
        epis_data = [['EPI', 'Quantidade']]
        for epi in relatorio['top_epis_entregues']:
            epis_data.append([epi['nome'][:40], str(epi['total'])])
        
        epis_table = Table(epis_data, colWidths=[350, 120])
        epis_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.063, 0.725, 0.506)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(epis_table)
    
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("Documento gerado pelo GestorEPI - Painel Master", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=relatorio_{empresa.get("nome", "empresa")}_{datetime.now().strftime("%Y%m%d")}.pdf'}
    )

@api_router.get('/master/empresas/{empresa_id}/export/excel')
async def export_empresa_relatorio_excel(
    empresa_id: str,
    periodo: int = 30,
    current_user: dict = Depends(require_super_admin())
):
    """Exporta relatório da empresa em Excel"""
    db = await get_db()
    
    empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    relatorio = await get_empresa_relatorio(empresa_id, periodo, current_user)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    
    # Estilos
    header_fill = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Título
    ws['A1'] = f"Relatório - {empresa.get('nome')}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Período: {periodo} dias"
    
    # Informações da Empresa
    ws['A4'] = "INFORMAÇÕES DA EMPRESA"
    ws['A4'].font = Font(bold=True, size=12)
    info = [
        ('CNPJ', empresa.get('cnpj', '-')),
        ('Plano', empresa.get('plano', '-')),
        ('Status', empresa.get('status', '-')),
        ('Responsável', empresa.get('responsavel', '-')),
        ('E-mail', empresa.get('email', '-')),
    ]
    for i, (label, value) in enumerate(info, 5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
    
    # Métricas
    row = 12
    ws[f'A{row}'] = "MÉTRICAS DE USO"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    metrics = [
        ('Colaboradores Ativos', relatorio['colaboradores_ativos']),
        ('Colaboradores Total', relatorio['total_colaboradores']),
        ('Limite do Plano', relatorio['limite_colaboradores']),
        ('Uso do Plano (%)', f"{relatorio['uso_percentual']}%"),
        ('EPIs Cadastrados', relatorio['total_epis']),
        ('EPIs Estoque Baixo', relatorio['epis_estoque_baixo']),
        ('EPIs Vencendo', relatorio['epis_vencendo']),
        ('Entregas no Período', relatorio['entregas_periodo']),
        ('Devoluções no Período', relatorio['devolucoes_periodo']),
        ('Usuários', relatorio['total_usuarios']),
        ('Kits', relatorio['total_kits']),
    ]
    
    for label, value in metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=relatorio_{empresa.get("nome", "empresa")}_{datetime.now().strftime("%Y%m%d")}.xlsx'}
    )

# ===================== FASE 4: CONTROLE DE PLANOS AVANÇADO =====================

@api_router.get('/master/empresas/{empresa_id}/historico-planos')
async def get_historico_planos(empresa_id: str, current_user: dict = Depends(require_super_admin())):
    """Lista histórico de mudanças de plano de uma empresa"""
    db = await get_db()
    
    historico = await db.historico_planos.find({"empresa_id": empresa_id}).sort("alterado_em", -1).to_list(100)
    
    return [{
        "id": str(h['_id']),
        "plano_anterior": h.get('plano_anterior'),
        "plano_novo": h.get('plano_novo'),
        "limite_anterior": h.get('limite_anterior'),
        "limite_novo": h.get('limite_novo'),
        "motivo": h.get('motivo'),
        "alterado_por": h.get('alterado_por'),
        "alterado_em": h.get('alterado_em')
    } for h in historico]

@api_router.patch('/master/empresas/{empresa_id}/plano')
async def atualizar_plano_empresa(
    empresa_id: str,
    plano: str,
    limite: int,
    motivo: Optional[str] = None,
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    current_user: dict = Depends(require_super_admin())
):
    """Atualiza plano de uma empresa com registro de histórico"""
    db = await get_db()
    
    empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    # Registrar histórico
    historico = {
        "empresa_id": empresa_id,
        "plano_anterior": empresa.get('plano'),
        "plano_novo": plano,
        "limite_anterior": empresa.get('limite_colaboradores'),
        "limite_novo": limite,
        "motivo": motivo,
        "alterado_por": current_user.get('username'),
        "alterado_em": datetime.now(timezone.utc)
    }
    await db.historico_planos.insert_one(historico)
    
    # Atualizar empresa
    update_data = {
        "plano": plano,
        "limite_colaboradores": limite,
        "updated_at": datetime.now(timezone.utc)
    }
    
    if data_inicio:
        update_data["data_inicio_plano"] = datetime.fromisoformat(data_inicio.replace('Z', '+00:00'))
    if data_fim:
        update_data["data_fim_plano"] = datetime.fromisoformat(data_fim.replace('Z', '+00:00'))
    
    await db.empresas.update_one({"_id": ObjectId(empresa_id)}, {"$set": update_data})
    
    return {"message": "Plano atualizado com sucesso", "historico_id": str(historico.get('_id', ''))}

@api_router.get('/master/empresas/{empresa_id}/vigencia')
async def get_vigencia_plano(empresa_id: str, current_user: dict = Depends(require_super_admin())):
    """Retorna informações de vigência do plano"""
    db = await get_db()
    
    empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    data_fim = empresa.get('data_fim_plano')
    status_vigencia = "ativo"
    dias_restantes = None
    
    if data_fim:
        if data_fim.tzinfo is None:
            data_fim = data_fim.replace(tzinfo=timezone.utc)
        
        now = datetime.now(timezone.utc)
        delta = data_fim - now
        dias_restantes = delta.days
        
        if dias_restantes < 0:
            status_vigencia = "expirado"
        elif dias_restantes <= 30:
            status_vigencia = "expirando"
    
    return {
        "empresa_id": empresa_id,
        "plano": empresa.get('plano'),
        "data_inicio": empresa.get('data_inicio_plano'),
        "data_fim": empresa.get('data_fim_plano'),
        "dias_restantes": dias_restantes,
        "status_vigencia": status_vigencia
    }

# ===================== BACKUP DO SISTEMA =====================

BACKUP_DIR = ROOT_DIR / 'backups'
BACKUP_DIR.mkdir(exist_ok=True)

def format_size(size_bytes):
    """Formata tamanho em bytes para formato legível"""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024
    return f"{size_bytes:.1f} TB"

@api_router.post('/master/backup')
async def criar_backup(
    descricao: Optional[str] = None,
    current_user: dict = Depends(require_super_admin())
):
    """Cria backup manual do banco de dados"""
    db = await get_db()
    
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    backup_filename = f"backup_{timestamp}.json"
    backup_path = BACKUP_DIR / backup_filename
    
    try:
        # Coleções para backup
        colecoes = [
            'empresas', 'users', 'employees', 'epis', 'kits', 
            'deliveries', 'suppliers', 'companies', 'facial_templates',
            'stock_movements', 'historico_planos', 'biometric_consent_logs',
            'ficha_authentications'
        ]
        
        backup_data = {
            "metadata": {
                "criado_em": datetime.now(timezone.utc).isoformat(),
                "criado_por": current_user.get('username'),
                "descricao": descricao,
                "versao": "1.0",
                "colecoes": colecoes
            },
            "dados": {}
        }
        
        # Exportar cada coleção
        for colecao in colecoes:
            try:
                docs = await db[colecao].find({}).to_list(50000)
                # Converter ObjectId para string
                for doc in docs:
                    doc['_id'] = str(doc['_id'])
                    for key, value in doc.items():
                        if isinstance(value, ObjectId):
                            doc[key] = str(value)
                        elif isinstance(value, datetime):
                            doc[key] = value.isoformat()
                backup_data["dados"][colecao] = docs
            except Exception as e:
                logger.warning(f"Erro ao exportar coleção {colecao}: {e}")
                backup_data["dados"][colecao] = []
        
        # Salvar arquivo
        with open(backup_path, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, ensure_ascii=False, indent=2, default=str)
        
        # Tamanho do arquivo
        file_size = backup_path.stat().st_size
        
        # Registrar backup no banco
        backup_record = {
            "nome_arquivo": backup_filename,
            "caminho": str(backup_path),
            "tamanho_bytes": file_size,
            "colecoes": colecoes,
            "descricao": descricao,
            "criado_por": current_user.get('username'),
            "criado_em": datetime.now(timezone.utc),
            "status": "completed"
        }
        result = await db.backups.insert_one(backup_record)
        
        # Limpeza de backups antigos (manter últimos 7 dias)
        await limpar_backups_antigos(db)
        
        return {
            "id": str(result.inserted_id),
            "nome_arquivo": backup_filename,
            "tamanho_bytes": file_size,
            "tamanho_formatado": format_size(file_size),
            "colecoes_incluidas": colecoes,
            "criado_em": backup_record["criado_em"].isoformat(),
            "status": "completed",
            "message": "Backup criado com sucesso"
        }
        
    except Exception as e:
        logger.error(f"Erro ao criar backup: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao criar backup: {str(e)}")

async def limpar_backups_antigos(db):
    """Remove backups com mais de 7 dias"""
    try:
        data_limite = datetime.now(timezone.utc) - timedelta(days=7)
        
        # Buscar backups antigos
        backups_antigos = await db.backups.find({
            "criado_em": {"$lt": data_limite}
        }).to_list(100)
        
        for backup in backups_antigos:
            # Remover arquivo
            try:
                caminho = Path(backup.get('caminho', ''))
                if caminho.exists():
                    caminho.unlink()
            except Exception as e:
                logger.warning(f"Erro ao remover arquivo de backup: {e}")
            
            # Remover registro
            await db.backups.delete_one({"_id": backup['_id']})
        
        logger.info(f"Removidos {len(backups_antigos)} backups antigos")
    except Exception as e:
        logger.error(f"Erro ao limpar backups antigos: {e}")

@api_router.get('/master/backups')
async def listar_backups(current_user: dict = Depends(require_super_admin())):
    """Lista todos os backups disponíveis"""
    db = await get_db()
    
    backups = await db.backups.find({}).sort("criado_em", -1).to_list(50)
    
    total_size = 0
    result = []
    
    for b in backups:
        size = b.get('tamanho_bytes', 0)
        total_size += size
        result.append({
            "id": str(b['_id']),
            "nome_arquivo": b.get('nome_arquivo'),
            "tamanho_bytes": size,
            "tamanho_formatado": format_size(size),
            "colecoes_incluidas": b.get('colecoes', []),
            "descricao": b.get('descricao'),
            "criado_por": b.get('criado_por'),
            "criado_em": b.get('criado_em'),
            "status": b.get('status', 'completed')
        })
    
    return {
        "backups": result,
        "total": len(result),
        "espaco_total_usado": format_size(total_size)
    }

@api_router.get('/master/backups/{backup_id}/download')
async def download_backup(backup_id: str, current_user: dict = Depends(require_super_admin())):
    """Download de um backup específico"""
    db = await get_db()
    
    backup = await db.backups.find_one({"_id": ObjectId(backup_id)})
    if not backup:
        raise HTTPException(status_code=404, detail="Backup não encontrado")
    
    caminho = Path(backup.get('caminho', ''))
    if not caminho.exists():
        raise HTTPException(status_code=404, detail="Arquivo de backup não encontrado")
    
    return StreamingResponse(
        open(caminho, 'rb'),
        media_type='application/json',
        headers={'Content-Disposition': f'attachment; filename={backup.get("nome_arquivo")}'}
    )

@api_router.delete('/master/backups/{backup_id}')
async def excluir_backup(backup_id: str, current_user: dict = Depends(require_super_admin())):
    """Exclui um backup específico"""
    db = await get_db()
    
    backup = await db.backups.find_one({"_id": ObjectId(backup_id)})
    if not backup:
        raise HTTPException(status_code=404, detail="Backup não encontrado")
    
    # Remover arquivo
    try:
        caminho = Path(backup.get('caminho', ''))
        if caminho.exists():
            caminho.unlink()
    except Exception as e:
        logger.warning(f"Erro ao remover arquivo: {e}")
    
    # Remover registro
    await db.backups.delete_one({"_id": ObjectId(backup_id)})
    
    return {"message": "Backup excluído com sucesso"}

# ===================== LGPD - CONFORMIDADE E PROTEÇÃO DE DADOS =====================

@api_router.get('/lgpd/dashboard')
async def get_lgpd_dashboard(current_user: dict = Depends(require_role('admin', 'seguranca_trabalho'))):
    """Dashboard de conformidade LGPD da empresa"""
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    empresa_filter = get_empresa_filter(current_user)
    
    # Colaboradores da empresa
    emp_query = {}
    if empresa_filter:
        emp_query.update(empresa_filter)
    total_colaboradores = await db.employees.count_documents(emp_query)
    
    # Colaboradores com biometria
    employees = await db.employees.find(emp_query).to_list(10000)
    employee_ids = [str(e['_id']) for e in employees]
    
    colaboradores_com_biometria = await db.facial_templates.count_documents({
        "employee_id": {"$in": employee_ids}
    })
    
    # Consentimentos registrados
    consent_query = {}
    if empresa_filter:
        consent_query.update(empresa_filter)
    total_consentimentos = await db.biometric_consent_logs.count_documents(consent_query)
    
    # Consentimentos por tipo
    consent_pipeline = [
        {"$match": consent_query},
        {"$group": {
            "_id": "$consent_type",
            "count": {"$sum": 1}
        }}
    ]
    consent_by_type = await db.biometric_consent_logs.aggregate(consent_pipeline).to_list(10)
    consent_map = {c['_id']: c['count'] for c in consent_by_type}
    
    # Solicitações de exclusão pendentes (se houver)
    exclusao_query = {"status": "pending"}
    if empresa_filter:
        exclusao_query.update(empresa_filter)
    exclusoes_pendentes = await db.lgpd_requests.count_documents(exclusao_query) if 'lgpd_requests' in await db.list_collection_names() else 0
    
    # Colaboradores SEM consentimento biométrico registrado
    colaboradores_sem_consentimento = []
    for emp in employees:
        emp_id = str(emp['_id'])
        has_consent = await db.biometric_consent_logs.find_one({
            "employee_id": emp_id,
            "consent_type": {"$in": ["granted", "initial"]}
        })
        if not has_consent:
            # Verificar se tem template facial (ou seja, biometria cadastrada)
            has_template = await db.facial_templates.find_one({"employee_id": emp_id})
            if has_template:
                colaboradores_sem_consentimento.append({
                    "id": emp_id,
                    "nome": emp.get('full_name'),
                    "cpf": emp.get('cpf'),
                    "departamento": emp.get('department')
                })
    
    return {
        "total_colaboradores": total_colaboradores,
        "colaboradores_com_biometria": colaboradores_com_biometria,
        "percentual_biometria": round((colaboradores_com_biometria / total_colaboradores * 100) if total_colaboradores > 0 else 0, 1),
        "total_consentimentos": total_consentimentos,
        "consentimentos_por_tipo": {
            "concedidos": consent_map.get("granted", 0) + consent_map.get("initial", 0),
            "revogados": consent_map.get("revoked", 0),
            "atualizados": consent_map.get("updated", 0)
        },
        "exclusoes_pendentes": exclusoes_pendentes,
        "colaboradores_sem_consentimento": colaboradores_sem_consentimento[:20],  # Limitar a 20
        "total_sem_consentimento": len(colaboradores_sem_consentimento),
        "status_conformidade": "conforme" if len(colaboradores_sem_consentimento) == 0 else "pendente"
    }

@api_router.get('/lgpd/consentimentos')
async def get_lgpd_consentimentos(
    current_user: dict = Depends(require_role('admin', 'seguranca_trabalho')),
    page: int = 1,
    limit: int = 50
):
    """Lista todos os consentimentos biométricos registrados"""
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    empresa_filter = get_empresa_filter(current_user)
    query = {}
    if empresa_filter:
        query.update(empresa_filter)
    
    skip = (page - 1) * limit
    total = await db.biometric_consent_logs.count_documents(query)
    
    consentimentos = await db.biometric_consent_logs.find(query).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Buscar nomes dos colaboradores
    emp_ids = list(set([c.get('employee_id') for c in consentimentos if c.get('employee_id')]))
    employees = await db.employees.find({"_id": {"$in": [ObjectId(e) for e in emp_ids]}}).to_list(len(emp_ids))
    emp_map = {str(e['_id']): e.get('full_name') for e in employees}
    
    return {
        "consentimentos": [{
            "id": str(c['_id']),
            "employee_id": c.get('employee_id'),
            "employee_name": emp_map.get(c.get('employee_id'), 'N/A'),
            "consent_type": c.get('consent_type'),
            "ip_address": c.get('ip_address'),
            "user_agent": c.get('user_agent', '')[:100],
            "created_at": c.get('created_at'),
            "granted_by": c.get('granted_by')
        } for c in consentimentos],
        "total": total,
        "page": page,
        "total_pages": (total + limit - 1) // limit
    }

@api_router.post('/lgpd/solicitar-exclusao/{employee_id}')
async def solicitar_exclusao_dados(
    employee_id: str,
    motivo: str = "Solicitação do titular",
    current_user: dict = Depends(require_role('admin'))
):
    """Solicita exclusão de dados de um colaborador (LGPD - direito ao esquecimento)"""
    db = await get_db()
    
    # MULTI-TENANT: Verificar se colaborador pertence à empresa
    empresa_filter = get_empresa_filter(current_user)
    emp_query = {"_id": ObjectId(employee_id)}
    if empresa_filter:
        emp_query.update(empresa_filter)
    
    employee = await db.employees.find_one(emp_query)
    if not employee:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    
    # Criar solicitação de exclusão
    request = {
        "employee_id": employee_id,
        "employee_name": employee.get('full_name'),
        "empresa_id": current_user.get('empresa_id'),
        "motivo": motivo,
        "solicitado_por": current_user.get('username'),
        "status": "pending",
        "created_at": datetime.now(timezone.utc)
    }
    
    result = await db.lgpd_requests.insert_one(request)
    
    return {
        "message": "Solicitação de exclusão registrada",
        "request_id": str(result.inserted_id),
        "status": "pending"
    }

@api_router.post('/lgpd/executar-exclusao/{employee_id}')
async def executar_exclusao_dados(
    employee_id: str,
    confirmar: bool = False,
    current_user: dict = Depends(require_role('admin'))
):
    """Executa a exclusão de dados biométricos e pessoais de um colaborador"""
    db = await get_db()
    
    if not confirmar:
        raise HTTPException(status_code=400, detail="É necessário confirmar a exclusão")
    
    # MULTI-TENANT: Verificar se colaborador pertence à empresa
    empresa_filter = get_empresa_filter(current_user)
    emp_query = {"_id": ObjectId(employee_id)}
    if empresa_filter:
        emp_query.update(empresa_filter)
    
    employee = await db.employees.find_one(emp_query)
    if not employee:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    
    dados_removidos = {
        "templates_faciais": 0,
        "consentimentos": 0,
        "foto_removida": False
    }
    
    # 1. Remover templates faciais
    result = await db.facial_templates.delete_many({"employee_id": employee_id})
    dados_removidos["templates_faciais"] = result.deleted_count
    
    # 2. Anonimizar consentimentos (manter log mas remover dados identificáveis)
    result = await db.biometric_consent_logs.update_many(
        {"employee_id": employee_id},
        {"$set": {
            "anonimizado": True,
            "ip_address": "ANONIMIZADO",
            "user_agent": "ANONIMIZADO",
            "anonimizado_em": datetime.now(timezone.utc),
            "anonimizado_por": current_user.get('username')
        }}
    )
    dados_removidos["consentimentos"] = result.modified_count
    
    # 3. Remover foto do colaborador
    if employee.get('photo_url'):
        # Limpar URL da foto no registro
        await db.employees.update_one(
            {"_id": ObjectId(employee_id)},
            {"$set": {"photo_url": None, "biometric_consent": False}}
        )
        dados_removidos["foto_removida"] = True
    
    # 4. Registrar log de exclusão LGPD
    await db.lgpd_exclusion_logs.insert_one({
        "employee_id": employee_id,
        "employee_name": employee.get('full_name'),
        "empresa_id": current_user.get('empresa_id'),
        "dados_removidos": dados_removidos,
        "executado_por": current_user.get('username'),
        "executado_em": datetime.now(timezone.utc)
    })
    
    # 5. Atualizar solicitação pendente se houver
    await db.lgpd_requests.update_many(
        {"employee_id": employee_id, "status": "pending"},
        {"$set": {"status": "completed", "completed_at": datetime.now(timezone.utc)}}
    )
    
    return {
        "message": f"Dados biométricos do colaborador {employee.get('full_name')} removidos com sucesso",
        "dados_removidos": dados_removidos
    }

@api_router.get('/lgpd/export-dados/{employee_id}')
async def export_dados_colaborador(
    employee_id: str,
    current_user: dict = Depends(require_role('admin'))
):
    """Exporta todos os dados de um colaborador (LGPD - direito à portabilidade)"""
    db = await get_db()
    
    # MULTI-TENANT: Verificar se colaborador pertence à empresa
    empresa_filter = get_empresa_filter(current_user)
    emp_query = {"_id": ObjectId(employee_id)}
    if empresa_filter:
        emp_query.update(empresa_filter)
    
    employee = await db.employees.find_one(emp_query)
    if not employee:
        raise HTTPException(status_code=404, detail="Colaborador não encontrado")
    
    # Coletar todos os dados do colaborador
    dados = {
        "dados_pessoais": {
            "nome": employee.get('full_name'),
            "cpf": employee.get('cpf'),
            "rg": employee.get('rg'),
            "email": employee.get('email'),
            "telefone": employee.get('phone'),
            "data_nascimento": str(employee.get('birth_date')) if employee.get('birth_date') else None,
            "departamento": employee.get('department'),
            "cargo": employee.get('position'),
            "data_admissao": str(employee.get('admission_date')) if employee.get('admission_date') else None,
            "matricula": employee.get('registration_number'),
            "status": employee.get('status')
        },
        "consentimentos_biometricos": [],
        "entregas_epi": [],
        "biometria_cadastrada": False
    }
    
    # Consentimentos
    consentimentos = await db.biometric_consent_logs.find({"employee_id": employee_id}).to_list(100)
    dados["consentimentos_biometricos"] = [{
        "tipo": c.get('consent_type'),
        "data": str(c.get('created_at')),
        "ip": c.get('ip_address'),
        "concedido_por": c.get('granted_by')
    } for c in consentimentos]
    
    # Verificar biometria
    template = await db.facial_templates.find_one({"employee_id": employee_id})
    dados["biometria_cadastrada"] = template is not None
    
    # Entregas de EPI
    entregas = await db.deliveries.find({"employee_id": employee_id}).sort("created_at", -1).to_list(500)
    dados["entregas_epi"] = [{
        "data": str(e.get('created_at')),
        "tipo": "devolução" if e.get('is_return') else "entrega",
        "itens": [{
            "epi": i.get('epi_name'),
            "quantidade": i.get('quantity'),
            "ca": i.get('ca_number')
        } for i in e.get('items', [])],
        "responsavel": e.get('delivered_by_name')
    } for e in entregas]
    
    dados["exportado_em"] = datetime.now(timezone.utc).isoformat()
    dados["exportado_por"] = current_user.get('username')
    
    # Retornar como JSON para download
    return StreamingResponse(
        io.BytesIO(json.dumps(dados, ensure_ascii=False, indent=2, default=str).encode('utf-8')),
        media_type='application/json',
        headers={'Content-Disposition': f'attachment; filename=dados_colaborador_{employee_id}_{datetime.now().strftime("%Y%m%d")}.json'}
    )

# ===================== RELATÓRIOS AVANÇADOS COM GRÁFICOS =====================

@api_router.get('/relatorios/entregas-por-periodo')
async def get_entregas_por_periodo(
    current_user: dict = Depends(get_current_user),
    periodo: int = 30,
    agrupamento: str = "dia"  # dia, semana, mes
):
    """Retorna dados de entregas agrupados por período para gráficos"""
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    empresa_filter = get_empresa_filter(current_user)
    
    data_inicio = datetime.now(timezone.utc) - timedelta(days=periodo)
    
    match_query = {
        "is_return": False,
        "created_at": {"$gte": data_inicio}
    }
    if empresa_filter:
        match_query.update(empresa_filter)
    
    # Definir formato de agrupamento
    if agrupamento == "semana":
        date_format = "%Y-W%U"
    elif agrupamento == "mes":
        date_format = "%Y-%m"
    else:
        date_format = "%Y-%m-%d"
    
    pipeline = [
        {"$match": match_query},
        {"$group": {
            "_id": {"$dateToString": {"format": date_format, "date": "$created_at"}},
            "total_entregas": {"$sum": 1},
            "total_itens": {"$sum": {"$size": "$items"}}
        }},
        {"$sort": {"_id": 1}}
    ]
    
    resultados = await db.deliveries.aggregate(pipeline).to_list(100)
    
    # Também contar devoluções
    match_query["is_return"] = True
    pipeline_dev = [
        {"$match": match_query},
        {"$group": {
            "_id": {"$dateToString": {"format": date_format, "date": "$created_at"}},
            "total_devolucoes": {"$sum": 1}
        }},
        {"$sort": {"_id": 1}}
    ]
    
    devolucoes = await db.deliveries.aggregate(pipeline_dev).to_list(100)
    dev_map = {d['_id']: d['total_devolucoes'] for d in devolucoes}
    
    return {
        "periodo_dias": periodo,
        "agrupamento": agrupamento,
        "dados": [{
            "periodo": r['_id'],
            "entregas": r['total_entregas'],
            "itens": r['total_itens'],
            "devolucoes": dev_map.get(r['_id'], 0)
        } for r in resultados]
    }

@api_router.get('/relatorios/consumo-epis')
async def get_consumo_epis(
    current_user: dict = Depends(get_current_user),
    periodo: int = 90,
    limite: int = 10
):
    """Retorna análise de consumo de EPIs para gráficos"""
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    empresa_filter = get_empresa_filter(current_user)
    
    data_inicio = datetime.now(timezone.utc) - timedelta(days=periodo)
    
    match_query = {
        "is_return": False,
        "created_at": {"$gte": data_inicio}
    }
    if empresa_filter:
        match_query.update(empresa_filter)
    
    # Top EPIs mais consumidos
    pipeline = [
        {"$match": match_query},
        {"$unwind": "$items"},
        {"$group": {
            "_id": {"epi_id": "$items.epi_id", "epi_name": "$items.epi_name"},
            "total_quantidade": {"$sum": "$items.quantity"},
            "total_entregas": {"$sum": 1}
        }},
        {"$sort": {"total_quantidade": -1}},
        {"$limit": limite}
    ]
    
    top_epis = await db.deliveries.aggregate(pipeline).to_list(limite)
    
    # Consumo por departamento
    pipeline_dept = [
        {"$match": match_query},
        {"$lookup": {
            "from": "employees",
            "localField": "employee_id",
            "foreignField": "_id",
            "as": "employee_info"
        }},
        {"$unwind": {"path": "$employee_info", "preserveNullAndEmptyArrays": True}},
        {"$group": {
            "_id": "$employee_info.department",
            "total_entregas": {"$sum": 1},
            "total_itens": {"$sum": {"$size": "$items"}}
        }},
        {"$match": {"_id": {"$ne": None}}},
        {"$sort": {"total_itens": -1}},
        {"$limit": 10}
    ]
    
    consumo_dept = await db.deliveries.aggregate(pipeline_dept).to_list(10)
    
    return {
        "periodo_dias": periodo,
        "top_epis": [{
            "epi_id": e['_id']['epi_id'],
            "epi_name": e['_id']['epi_name'] or "N/A",
            "quantidade": e['total_quantidade'],
            "entregas": e['total_entregas']
        } for e in top_epis],
        "consumo_por_departamento": [{
            "departamento": c['_id'] or "Sem departamento",
            "entregas": c['total_entregas'],
            "itens": c['total_itens']
        } for c in consumo_dept]
    }

@api_router.get('/relatorios/estoque-critico')
async def get_estoque_critico(current_user: dict = Depends(get_current_user)):
    """Retorna EPIs com estoque crítico para dashboard"""
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    empresa_filter = get_empresa_filter(current_user)
    
    query = {"$expr": {"$lte": ["$current_stock", "$min_stock"]}}
    if empresa_filter:
        query.update(empresa_filter)
    
    epis_criticos = await db.epis.find(query).sort("current_stock", 1).to_list(50)
    
    # Calcular nível de criticidade
    result = []
    for epi in epis_criticos:
        current = epi.get('current_stock', 0)
        min_stock = epi.get('min_stock', 1)
        
        if min_stock > 0:
            percentual = (current / min_stock) * 100
        else:
            percentual = 0
        
        if current == 0:
            nivel = "zerado"
        elif percentual <= 25:
            nivel = "critico"
        elif percentual <= 50:
            nivel = "baixo"
        else:
            nivel = "atencao"
        
        result.append({
            "id": str(epi['_id']),
            "nome": epi.get('name'),
            "ca_number": epi.get('ca_number'),
            "estoque_atual": current,
            "estoque_minimo": min_stock,
            "percentual_estoque": round(percentual, 1),
            "nivel_criticidade": nivel
        })
    
    return {
        "total_criticos": len(result),
        "zerados": len([e for e in result if e['nivel_criticidade'] == 'zerado']),
        "criticos": len([e for e in result if e['nivel_criticidade'] == 'critico']),
        "baixos": len([e for e in result if e['nivel_criticidade'] == 'baixo']),
        "epis": result
    }

@api_router.get('/relatorios/vencimentos')
async def get_vencimentos(
    current_user: dict = Depends(get_current_user),
    dias: int = 90
):
    """Retorna EPIs com validade próxima do vencimento"""
    db = await get_db()
    
    # MULTI-TENANT: Filtrar por empresa
    empresa_filter = get_empresa_filter(current_user)
    
    data_limite = datetime.now(timezone.utc) + timedelta(days=dias)
    
    query = {
        "$or": [
            {"validity_date": {"$ne": None, "$lte": data_limite}},
            {"ca_validity": {"$ne": None, "$lte": data_limite}}
        ]
    }
    if empresa_filter:
        query.update(empresa_filter)
    
    epis_vencendo = await db.epis.find(query).to_list(100)
    
    now = datetime.now(timezone.utc)
    result = []
    
    for epi in epis_vencendo:
        validity = epi.get('validity_date') or epi.get('ca_validity')
        if validity:
            if validity.tzinfo is None:
                validity = validity.replace(tzinfo=timezone.utc)
            
            dias_restantes = (validity - now).days
            
            if dias_restantes < 0:
                status = "vencido"
            elif dias_restantes <= 7:
                status = "critico"
            elif dias_restantes <= 30:
                status = "urgente"
            else:
                status = "atencao"
            
            result.append({
                "id": str(epi['_id']),
                "nome": epi.get('name'),
                "ca_number": epi.get('ca_number'),
                "data_vencimento": validity.isoformat(),
                "dias_restantes": dias_restantes,
                "status": status,
                "tipo_validade": "produto" if epi.get('validity_date') else "CA"
            })
    
    # Ordenar por dias restantes
    result.sort(key=lambda x: x['dias_restantes'])
    
    return {
        "periodo_dias": dias,
        "total": len(result),
        "vencidos": len([e for e in result if e['status'] == 'vencido']),
        "criticos": len([e for e in result if e['status'] == 'critico']),
        "urgentes": len([e for e in result if e['status'] == 'urgente']),
        "epis": result
    }

app.include_router(api_router)
